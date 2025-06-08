# -*- coding: utf-8 -*-
import sys  # 添加这行
import os
import re
import math
import json
import logging
import shutil
import sqlite3
import tempfile
import traceback
import subprocess
import datetime
import weakref
from pathlib import Path
from copy import copy, deepcopy
from datetime import datetime
# ======================= 第三方库导入 =======================
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PIL import Image  # 确保导入PIL

# ======================= PyQt5核心导入 =======================
from PyQt5.QtCore import (
    Qt, QSize, QRect, QPoint, QUrl, QMimeData,
    QDate, QSettings, QTimer, pyqtSignal, QByteArray  # 添加QByteArray
)
from PyQt5.QtGui import (
    QIcon, QColor, QPainter, QPen, QPixmap, QFont,
    QDesktopServices, QDrag, QTextCursor, QFontInfo, QFontDatabase  # 添加QFontDatabase
)
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QDialog, QWidget,
    QTableWidget, QTableWidgetItem, QVBoxLayout, QHBoxLayout,
    QFormLayout, QPushButton, QMessageBox, QLabel, QLineEdit,
    QDateEdit, QFileDialog, QTextEdit, QComboBox, QFrame,
    QMenu, QAction, QDialogButtonBox, QStyledItemDelegate,
    QCompleter, QHeaderView, QGraphicsView, QGraphicsScene,
    QGraphicsTextItem, QGraphicsItem, QSpinBox, QListWidget,
    QScrollArea, QGroupBox, QColorDialog, QCheckBox,
    QFontComboBox, QListWidgetItem, QAbstractItemView,
    QRadioButton, QButtonGroup, QToolBar, QSizePolicy, QFontDialog, QProgressDialog,
    QInputDialog  # 添加QInputDialog
)
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog

# 屏蔽 DeprecationWarning
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

# ======================= 定义全局 field_manager 变量 =======================
field_manager = None

def check_dependencies():
    try:
        import reportlab
        from PIL import Image
        return True
    except ImportError:
        return False
    
# 在启动时检查依赖
if not check_dependencies():
    QMessageBox.critical(
        None, 
        "环境错误", 
        "缺少打印组件依赖，请安装：\n\npip install reportlab Pillow"
    )
    sys.exit(1)    

# ======================= 环境配置 =======================
os.environ["QT_LOGGING_RULES"] = "qt.png.warning=false"
CONFIG_FILE = "print_config.json"

class PrintDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.mappings = {}  # 存储字段映射关系
        self.template_path = ""  # 添加缺失的属性
        self.copies_spinbox = QSpinBox()  # 添加缺失的控件
        self.init_ui()  # 添加初始化UI的方法
        self.load_config()  # 初始化时加载配置
        
    def init_ui(self):
        """初始化UI界面"""
        layout = QVBoxLayout()
        # 添加模板路径选择控件
        self.template_combo = QComboBox()
        layout.addWidget(self.template_combo)
        
        # 添加打印份数控件
        copies_layout = QHBoxLayout()
        copies_layout.addWidget(QLabel("打印份数:"))
        self.copies_spinbox.setRange(1, 100)
        copies_layout.addWidget(self.copies_spinbox)
        layout.addLayout(copies_layout)
        
        # 添加按钮
        btn_save = QPushButton("保存配置")
        btn_save.clicked.connect(self.save_config)
        layout.addWidget(btn_save)
        
        self.setLayout(layout)

    def load_config(self):
        """加载已有配置"""
        if Path(CONFIG_FILE).exists():
            try:
                with open(CONFIG_FILE) as f:
                    config = json.load(f)
                    self.template_path = config.get("template_path", "")
                    self.mappings = config.get("default_mappings", {})
                    # 自动填充界面控件...
            except:
                self.mappings = {}  # 配置文件损坏时重置
# --------------------------- 工具函数 ---------------------------
def resource_path(relative_path):
    """获取资源的绝对路径（增强字体路径处理）"""
    base_path = os.path.abspath(os.path.dirname(__file__))
    
    # 优先检查程序目录下的fonts文件夹
    font_path = os.path.join(base_path, "fonts", os.path.basename(relative_path))
    if os.path.exists(font_path):
        return font_path

    # 备用路径处理（原有逻辑）
    path = os.path.join(base_path, *relative_path.split('/'))
    if not os.path.exists(path):
        doc_path = os.path.join(os.path.expanduser("~"), "Documents", "人事档案系统", "templates")
        os.makedirs(doc_path, exist_ok=True)
        path = os.path.join(doc_path, os.path.basename(relative_path))
        
    return path.replace("/", os.sep)

def get_db_path():
    """数据库路径应固定在用户文档目录"""
    doc_path = os.path.expanduser("~/Documents/人事档案系统")
    os.makedirs(doc_path, exist_ok=True)
    return os.path.join(doc_path, "archive.db")

def validate_id_number(id_number):
    """验证中国大陆身份证号格式"""
    if not isinstance(id_number, str) or len(id_number) != 18:
        return False
    try:
        # 验证校验码
        factors = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
        check_codes = ['1', '0', 'X', '9', '8', '7', '6', '5', '4', '3', '2']
        total = sum(int(a) * b for a, b in zip(id_number[:17], factors))
        return id_number[-1].upper() == check_codes[total % 11]
    except:
        return False
def init_logger():
    """初始化日志系统（增强健壮性）"""
    try:
        logger = logging.getLogger("archive_manager")
        logger.setLevel(logging.DEBUG)
        
        # 确保只初始化一次
        if logger.hasHandlers():
            return logger
            
        # 日志目录处理
        log_dir = os.path.join(os.path.expanduser("~"), "Documents", "人事档案系统")
        try:
            os.makedirs(log_dir, exist_ok=True)
        except PermissionError:
            log_dir = tempfile.gettempdir()
            logger.warning(f"使用临时目录: {log_dir}")
        
        # 文件处理器
        file_handler = logging.FileHandler(
            os.path.join(log_dir, "archive_manager.log"), 
            encoding='utf-8'
        )
        file_handler.setFormatter(
            logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        )
        
        # 控制台处理器
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(file_handler.formatter)
        
        logger.addHandler(file_handler)
        logger.addHandler(console_handler)
        
        return logger
    except Exception as e:
        print(f"致命错误：无法初始化日志系统 {str(e)}")
        sys.exit(1)

# 确保在最顶层初始化logger
logger = init_logger()  

# --------------------------- 字段管理类 ---------------------------
class FieldManager:
    def __init__(self):
        logger.debug("字段管理器初始化开始")
        # 延迟初始化验证
        if not hasattr(logger, 'handlers'):
            logger.error("字段管理器初始化时日志系统未就绪")
            raise RuntimeError("字段管理器初始化时日志系统未就绪")
            
        self.logger = logger
        self.logger.debug("字段管理器初始化中...")
        self.fields = self.load_fields()
        self.logger.debug(f"字段配置加载完成，共 {len(self.fields)} 个字段")
        
    def validate_fields(self, fields):
        """验证字段配置"""
        if not isinstance(fields, list):
            raise ValueError("字段配置必须是列表")
            
        names = set()
        for field in fields:
            if not isinstance(field, dict):
                raise ValueError("每个字段必须是字典")
            if 'name' not in field:
                raise ValueError("字段缺少'name'属性")
            if not isinstance(field['name'], str) or not field['name'].strip():
                raise ValueError("字段名必须是非空字符串")
            if field['name'] in names:
                raise ValueError(f"重复字段名: {field['name']}")
            names.add(field['name'])
    
    def get_create_table_sql(self):
        """生成建表SQL语句"""
        columns = ['id INTEGER PRIMARY KEY AUTOINCREMENT']
        for field in self.fields:
            col_def = f'"{field["name"]}" TEXT'
            if field.get('unique'):
                col_def += ' UNIQUE'
            if field.get('required'):
                col_def += ' NOT NULL'
            columns.append(col_def)
        return f'CREATE TABLE IF NOT EXISTS personnel ({", ".join(columns)})'
    
    def get_field_names(self):
        """获取所有字段名"""
        return [field['name'] for field in self.fields]

    def get_default_fields(self):
        """返回完整的默认字段列表"""
        return [    
            {'name': '档案编号', 'type': 'str', 'required': True, 'unique': True},
            {'name': '姓名', 'type': 'str', 'required': True},
            {'name': '身份证号', 'type': 'str', 'required': True, 'unique': True},
            {'name': '身份', 'type': 'str', 'required': False},
            {'name': '籍贯', 'type': 'str', 'required': False},
            {'name': '一级单位', 'type': 'str', 'required': False},
            {'name': '二级单位', 'type': 'str', 'required': False},
            {'name': '出生日期', 'type': 'date', 'required': False},
            {'name': '参加工作时间', 'type': 'date', 'required': False},
            {'name': '入党日期', 'type': 'date', 'required': False},
            {'name': '工作经历', 'type': 'text', 'required': False},
            {'name': '学历', 'type': 'str', 'required': False},
            {'name': '档案流转记录', 'type': 'text', 'required': False},
            {'name': '电子档案', 'type': 'str', 'required': False},
            {'name': '备注', 'type': 'text', 'required': False},
            {'name': '学习经历', 'type': 'text', 'required': False},
        ]

    def load_fields(self):
        """加载字段配置"""
        try:
            template_dir = resource_path('templates')
            os.makedirs(template_dir, exist_ok=True)
            path = os.path.join(template_dir, 'fields.json')
            
            # 如果配置文件不存在，创建默认配置
            if not os.path.exists(path):
                default_fields = self.get_default_fields()
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump({'fields': default_fields}, f, indent=2, ensure_ascii=False)
                return default_fields
                
            # 加载现有配置
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                loaded_fields = data.get('fields', self.get_default_fields())
                self.validate_fields(loaded_fields)
                return loaded_fields
                
        except Exception as e:
            self.logger.error(f"加载字段配置失败: {str(e)}")
            # 添加详细的错误信息
            self.logger.error(f"错误详情: {traceback.format_exc()}")
            return self.get_default_fields()


        
class FieldManagerDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.field_manager = field_manager
        self.init_ui()
        self.load_fields()

    def init_ui(self):
        layout = QVBoxLayout()
        
        # 字段表格
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["字段名称", "类型", "必填", "唯一", "操作"])
        self.table.horizontalHeader().setStretchLastSection(True)
        
        # 操作按钮
        btn_layout = QHBoxLayout()
        btn_add = QPushButton("添加字段")
        btn_add.clicked.connect(self.add_field)
        btn_save = QPushButton("保存配置")
        btn_save.clicked.connect(self.save_config)
        
        btn_layout.addWidget(btn_add)
        btn_layout.addWidget(btn_save)
        
        layout.addWidget(self.table)
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
        self.resize(800, 500)

    def load_fields(self):
        self.table.setRowCount(0)
     
        for idx, field in enumerate(self.field_manager.fields):
            self.table.insertRow(idx)
            
            # 字段名称
            name_item = QTableWidgetItem(field['name'])
            
            # 类型选择
            type_combo = QComboBox()
            type_combo.addItems(['str', 'date', 'text'])
            type_combo.setCurrentText(field.get('type', 'str'))
            
            # 必填
            required_check = QCheckBox()
            required_check.setChecked(field.get('required', False))
            
            # 唯一
            unique_check = QCheckBox()
            unique_check.setChecked(field.get('unique', False))
            
            # 删除按钮
            del_btn = QPushButton("删除")
            del_btn.clicked.connect(lambda _, row=idx: self.delete_field(row))
            
            self.table.setItem(idx, 0, name_item)
            self.table.setCellWidget(idx, 1, type_combo)
            self.table.setCellWidget(idx, 2, required_check)
            self.table.setCellWidget(idx, 3, unique_check)
            self.table.setCellWidget(idx, 4, del_btn)

    def add_field(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        name_item = QTableWidgetItem("新字段")
        type_combo = QComboBox()
        type_combo.addItems(['str', 'date', 'text'])
        
        self.table.setItem(row, 0, name_item)
        self.table.setCellWidget(row, 1, type_combo)
        self.table.setCellWidget(row, 2, QCheckBox())
        self.table.setCellWidget(row, 3, QCheckBox())
        self.table.setCellWidget(row, 4, QPushButton("删除"))

    def delete_field(self, row):
        """删除字段"""
        if row < 0 or row >= len(self.field_manager.fields):
            return
            
        # 从字段列表中删除
        self.field_manager.fields.pop(row)
        
        # 从表格中删除行
        self.table.removeRow(row)
        
        # 立即更新UI
        self.load_fields()

    def save_config(self):
        try:
            # 从表格收集字段数据
            fields = []
            seen_names = set()
            
            for row in range(self.table.rowCount()):
                name_item = self.table.item(row, 0)
                if not name_item or not name_item.text().strip():
                    continue
                    
                field_name = name_item.text().strip()
                if field_name in seen_names:
                    QMessageBox.warning(self, "错误", f"字段名'{field_name}'重复！")
                    return
                seen_names.add(field_name)
                
                type_combo = self.table.cellWidget(row, 1)
                required_check = self.table.cellWidget(row, 2)
                unique_check = self.table.cellWidget(row, 3)
                
                fields.append({
                    'name': field_name,
                    'type': type_combo.currentText(),
                    'required': required_check.isChecked(),
                    'unique': unique_check.isChecked()
                })
            
            # 保存到文件
            template_dir = resource_path('templates')
            os.makedirs(template_dir, exist_ok=True)
            save_path = os.path.join(template_dir, 'fields.json')
            
            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump({'fields': fields}, f, ensure_ascii=False, indent=2)
            
            # 更新配置并提示重启
            self.field_manager.fields = fields
            QMessageBox.information(
                self, "成功", 
                "字段配置已保存！\n请重启程序应用数据库变更。"
            )
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败: {str(e)}")
# --------------------------- 数据库操作 ---------------------------
class DBManager:
    _instance = None
    
    def __new__(cls):
        if not cls._instance:
            cls._instance = super().__new__(cls)
            cls._instance.fields = cls._instance.load_fields()
            cls._instance._init_db()
        return cls._instance
    
    def _init_db(self):
        """初始化数据库连接"""
        try:
            db_path = get_db_path()
            self.conn = sqlite3.connect(
                db_path,
                check_same_thread=False,
                timeout=30
            )
            self.conn.execute("PRAGMA journal_mode=WAL")
            self.conn.execute("PRAGMA foreign_keys=ON")
            logger.info(f"数据库连接已建立: {db_path}")
        except Exception as e:
            logger.critical(f"数据库连接失败: {str(e)}")
            raise
    
    def get_connection(self):
        """获取数据库连接"""
        if not hasattr(self, 'conn') or not self.conn:
            self._init_db()
        return self.conn
    
    def close(self):
        """关闭数据库连接"""
        if hasattr(self, 'conn') and self.conn:
            self.conn.close()
            self.conn = None

# ======================= 修改 init_database 函数 =======================
# ======================= 修改 init_database 函数 =======================
def init_database():
    """初始化数据库（使用全局 field_manager）"""
    global field_manager  # 声明使用全局变量
    
    # 确保field_manager已经初始化
    if field_manager is None:
        field_manager = FieldManager()  # 初始化全局字段管理器
    
    # 首先获取数据库路径，确保在异常处理中可用
    db_path = get_db_path()
    
    try:
        logger.info(f"初始化数据库路径: {db_path}")
        
        # 备份旧数据库（如果存在）
        if os.path.exists(db_path):
            backup_path = f"{db_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            shutil.copy2(db_path, backup_path)
            logger.info(f"已创建数据库备份: {backup_path}")
        
        # 重建数据库
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            
            # 获取当前字段配置
            create_sql = field_manager.get_create_table_sql()
            logger.debug(f"执行建表SQL: {create_sql}")
            
            cursor.execute("DROP TABLE IF EXISTS personnel")
            cursor.execute(create_sql)
            
            # 新增代码：获取预期的字段列表和数据库现有字段
            cursor.execute("PRAGMA table_info(personnel)")
            db_columns = [col[1] for col in cursor.fetchall()]  # 获取数据库现有字段
            expected_columns = set(field_manager.get_field_names())  # 获取配置字段
            
            # 修复字段验证逻辑
            existing_columns = set(db_columns[1:]) if db_columns else set()  # 跳过id列
            missing_columns = expected_columns - existing_columns
            
            if missing_columns:
                logger.warning(f"发现缺失字段：{missing_columns}，尝试自动迁移...")
                for col in missing_columns:
                    cursor.execute(f'ALTER TABLE personnel ADD COLUMN "{col}" TEXT')
                conn.commit()
            
    except Exception as e:
        logger.critical(f"数据库初始化失败: {str(e)}")
        error_msg = (
            f"数据库初始化失败！\n\n"
            f"错误详情: {str(e)}\n\n"
            f"建议操作:\n"
            f"1. 检查字段配置文件: {resource_path('templates/fields.json')}\n"
            f"2. 删除损坏的数据库文件: {db_path}\n"
            f"3. 重新启动程序"
        )
        QMessageBox.critical(None, "数据库错误", error_msg)
        raise RuntimeError(error_msg)
    
def update_personnel(db_path, original_id, data):
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            # 动态构建 SET 子句
            set_fields = [f'"{k}" = ?' for k in data.keys()]
            set_clause = ", ".join(set_fields)
            # 参数顺序：新值 + 原始身份证号
            params = list(data.values()) + [original_id]
            # 执行更新
            cursor.execute(
                f"UPDATE personnel SET {set_clause} WHERE 身份证号 = ?",
                params
            )
            conn.commit()  # 显式提交事务
            return cursor.rowcount > 0
    except sqlite3.IntegrityError as e:
        # 处理唯一性冲突
        logger.error(f"唯一性冲突: {str(e)}")
        return False

def get_all_personnel(db_path):
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM personnel")
        return cursor.fetchall()
    try:
        with sqlite3.connect(db_path, check_same_thread=False) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM personnel")
            raw_data = cursor.fetchall()
            
            # 补全字段
            columns = ['id', '档案编号', '姓名', '身份证号', '身份', '籍贯','一级单位', '二级单位',
                      '出生日期', '参加工作时间', '入党日期', '工作经历', '学历',
                      '档案流转记录', '电子档案', '备注']
            return [
                {col: row[col] if col in row.keys() else "" for col in columns}
                for row in raw_data
            ]
    except Exception as e:
        logger.error(f"获取数据失败: {str(e)}")
        return []
    
def add_personnel(db_path, data):
    """新增档案"""
    try:
        # 确保所有必填字段都有值
        required_fields = [f['name'] for f in field_manager.fields if f.get('required', False)]
        for field in required_fields:
            if not data.get(field, '').strip():
                raise ValueError(f"必填字段 '{field}' 不能为空")
        
        # 验证身份证号格式
        if '身份证号' in data and not validate_id_number(data['身份证号']):
            raise ValueError("身份证号格式错误（需18位数字）")
            
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            # 使用字段管理器中的字段顺序
            fields = field_manager.get_field_names()
            values = [data.get(field, '') for field in fields]
            
            # 动态构建SQL
            placeholders = ','.join(['?'] * len(fields))
            query = f'''
                INSERT INTO personnel ({",".join(fields)})
                VALUES ({placeholders})
            '''
            cursor.execute(query, values)
            conn.commit()
        logger.info(f"新增档案成功: {data.get('姓名', '')}（{data.get('身份证号', '')}）")
        return True
    except Exception as e:
        logger.error(f"新增失败: {str(e)}")
        QMessageBox.critical(None, "错误", f"新增失败: {str(e)}")  # 添加错误提示
        return False
     
def import_from_excel(db_path, file_path):
    # 获取所有字段名
    all_columns = field_manager.get_field_names()
    
    # 读取Excel时指定字段类型为字符串
    df = pd.read_excel(file_path, dtype=str, keep_default_na=False)
    
    # 填充缺失列
    for col in all_columns:
        if col not in df.columns:
            df[col] = ''
    
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            success_count = 0
            error_count = 0
            
            # 转义字段名（防止SQL注入）
            escaped_columns = [f'"{col}"' for col in all_columns]
            placeholders = ', '.join(['?'] * len(all_columns))
            
            for idx, row in df.iterrows():
                try:
                    # 构建插入语句
                    query = f'''
                        INSERT INTO personnel ({",".join(escaped_columns)})
                        VALUES ({placeholders})
                    '''
                    cursor.execute(query, tuple(str(row[col]) for col in all_columns))
                    success_count += 1
                except sqlite3.IntegrityError as e:
                    # 唯一性冲突跳过记录
                    error_count += 1
                    logger.warning(f"唯一性冲突跳过记录 {idx+1}: {str(e)}")
                except Exception as e:
                    error_count += 1
                    logger.error(f"导入记录 {idx+1} 失败: {str(e)}")
            
            conn.commit()
            
            # 记录导入结果
            logger.info(f"导入完成: 成功 {success_count} 条, 失败 {error_count} 条")
            return success_count
    except Exception as e:
        logger.error(f"导入失败：{str(e)}")
        raise

def delete_personnel(db_path, ids):
    """增强版删除方法"""
    try:
        if not ids:
            return False
            
        with sqlite3.connect(db_path, timeout=10) as conn:
            cursor = conn.cursor()
            
            # 构建安全查询
            placeholders = ','.join(['?'] * len(ids))
            query = f"DELETE FROM personnel WHERE 身份证号 IN ({placeholders})"
            
            cursor.execute(query, ids)
            if cursor.rowcount == 0:
                logger.warning(f"未找到匹配记录: {ids}")
                return False
                
            conn.commit()
            logger.info(f"删除成功: {cursor.rowcount}条记录")
            return True
    except sqlite3.OperationalError as e:
        error_msg = f"数据库操作失败: {str(e)}"
        if "locked" in str(e):
            error_msg += "\n请关闭其他可能正在使用数据库的程序"
        logger.error(error_msg)
        raise
    except Exception as e:
        logger.error(f"删除失败: {traceback.format_exc()}")
        raise
    
# 建议放在所有主界面类之后，保持代码结构清晰
class ConfigManager(QDialog):
    def __init__(self, mappings=None):  # 修改为可选参数
        super().__init__()
        self.mappings = mappings or {}  # 处理None情况
        self.setup_ui()
        self._populate_list()  # 初始化时加载配置
        
    def setup_ui(self):
        self.setWindowTitle("模板配置管理")
        layout = QVBoxLayout()
        
        # 列表控件
        self.list_widget = QListWidget()
        self._populate_list()

            # ----------------- 新增字段映射表 -----------------
        self.field_table = QTableWidget(0, 3)  # 初始化表格控件
        self.field_table.setHorizontalHeaderLabels(["模板字段", "映射字段", "示例值"])
        self.field_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.field_table)
            
        # 删除按钮
        btn_del = QPushButton("删除选中配置")
        btn_del.clicked.connect(self.delete_config)
        
        layout.addWidget(self.list_widget)
        layout.addWidget(btn_del)
        self.setLayout(layout)
    
    def _populate_list(self):
            self.list_widget.clear()
            config_path = os.path.join(resource_path('templates'), 'template_config.json')
            try:
                if os.path.exists(config_path):
                    with open(config_path, 'r', encoding='utf-8') as f:
                        self.mappings = json.load(f)
                else:
                    self.mappings = {}
                # 填充列表
                for template in self.mappings.keys():
                    self.list_widget.addItem(template)
            except Exception as e:
                logger.error(f"加载模板配置失败: {str(e)}")
    
    def delete_config(self):
        """删除选中配置"""
        selected = self.list_widget.currentRow()
        if selected >= 0:
            key = list(self.mappings.keys())[selected]
            del self.mappings[key]
            self._populate_list()

class MainWindow(QMainWindow):
    def contextMenuEvent(self, event):
        menu = QMenu()
        edit_action = QAction("编辑", self)
        delete_action = QAction("删除", self)
        export_action = QAction("导出选中", self)
        edit_action.triggered.connect(self.edit_selected_row)
        delete_action.triggered.connect(self.delete_selected)
        export_action.triggered.connect(self.export_selected)
        
        menu.addActions([edit_action, delete_action, export_action])
        menu.exec_(event.globalPos())
    pass
# --------------------------- 对话框类 ---------------------------
class DynamicFormDialog(QDialog):
    def __init__(self, parent=None, mode='add', row_data=None):
        super().__init__(parent)
        # 使用弱引用避免循环引用
        self.parent_ref = weakref.ref(parent) if parent else None
        self.mode = mode
        self.row_data = row_data or {}
        self.field_manager = field_manager
        self.widgets = {}
        self.setup_ui()
        self.setWindowTitle("新增档案" if mode == 'add' else "编辑档案")
        
        # 使用安全的方式设置焦点
        QTimer.singleShot(100, self.safe_focus_first_widget)

        
    def closeEvent(self, event):
        """安全关闭对话框"""
        try:
            # 断开所有信号连接
            for btn in self.findChildren(QPushButton):
                try:
                    btn.clicked.disconnect()
                except:
                    pass
            
            # 清除父窗口引用
            self.parent_ref = None
            super().closeEvent(event)
        except Exception as e:
            logger.error(f"关闭对话框异常: {str(e)}")
            event.accept()


    def showEvent(self, event):
        """对话框显示时强制获取焦点 - 安全版本"""
        super().showEvent(event)
        # 使用定时器确保安全设置焦点
        QTimer.singleShot(50, self.safe_focus_first_widget)

    def safe_focus_first_widget(self):
        """安全设置焦点到第一个控件"""
        try:
            if self.widgets:
                first_key = next(iter(self.widgets.keys()))
                first_widget = self.widgets[first_key]
                if first_widget and first_widget.isEnabled():
                    first_widget.setFocus()
        except Exception as e:
            logger.error(f"设置焦点失败: {str(e)}")

    def setup_ui(self):
        # 主滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        # 内容容器
        content = QWidget()
        layout = QFormLayout(content)
        layout.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        
        # 动态创建字段
        for field in self.field_manager.fields:
            label = QLabel(f"{field['name']}{'*' if field['required'] else ''}")
            widget = self._create_field_widget(field)
            layout.addRow(label, widget)
            self.widgets[field['name']] = widget

        scroll.setWidget(content)
        
        # 按钮区域 - 关键修改：使用QDialogButtonBox
        self.button_box = QDialogButtonBox(
            QDialogButtonBox.Save | QDialogButtonBox.Cancel,
            Qt.Horizontal, self
        )
        self.button_box.accepted.connect(self._on_save)
        self.button_box.rejected.connect(self.reject)
        # 新增代码：汉化按钮文本
        self.button_box.button(QDialogButtonBox.Save).setText("保存")
        self.button_box.button(QDialogButtonBox.Cancel).setText("取消")
            # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(scroll)
        main_layout.addWidget(self.button_box)  # 这里修正为self.button_box
        
        # 关键修复：使用定时器延迟设置焦点，避免闪退
        QTimer.singleShot(100, self.focus_first_widget)
        self.resize(800, 1000)  

    def focus_first_widget(self):
        """设置焦点到第一个控件"""
        if self.widgets:
            first_widget = next(iter(self.widgets.values()))
            first_widget.setFocus()

    def _create_field_widget(self, field):
        """创建字段对应的控件"""
        field_type = field.get('type', 'str')
        
        if field_type == 'date':
            widget = QDateEdit()
            widget.setDisplayFormat("yyyy-MM-dd")
            widget.setDate(QDate.currentDate())  # 设置为当前日期而不是固定值
            if self.mode == 'edit':
                date_str = self.row_data.get(field['name'], '')
                if date_str:
                    try:
                        widget.setDate(QDate.fromString(date_str, "yyyy-MM-dd"))
                    except:
                        widget.setDate(QDate.currentDate())
        elif field_type == 'text':
            widget = QTextEdit()
            widget.setPlainText(self.row_data.get(field['name'], ''))
        else:
            widget = QLineEdit()
            widget.setText(self.row_data.get(field['name'], ''))
        
        return widget


    def _on_cancel(self):
        """取消按钮处理 - 确保立即关闭"""
        self.reject()

    def closeEvent(self, event):
        """安全关闭对话框 - 简化版本"""
        try:
            # 不再断开所有信号，Qt会自动管理
            self.parent_ref = None  # 清除父对象引用
            super().closeEvent(event)
        except Exception as e:
            logger.error(f"关闭对话框异常: {str(e)}")
            event.accept()
    def _validate_data(self):
        try:
            data = {}
            missing_fields = []
            
            for field in self.field_manager.fields:
                widget = self.widgets[field['name']]
                value = self._get_widget_value(widget, field)
                
                if field.get('required', False) and not value.strip():
                    missing_fields.append(field['name'])
                data[field['name']] = value

            if missing_fields:
                raise ValueError(f"以下字段必填：{', '.join(missing_fields)}")

            if '身份证号' in data and not validate_id_number(data['身份证号']):
                raise ValueError("身份证号格式错误（需18位数字）")

            self._validated_data = data
            return True
            
        except ValueError as e:
            if missing_fields:
                field_name = missing_fields[0]
                self.widgets[field_name].setFocus()
                if isinstance(self.widgets[field_name], (QLineEdit, QDateEdit)):
                    self.widgets[field_name].selectAll()
                    
            QMessageBox.warning(self, "输入错误", str(e))
            return False
        except Exception as e:
            logger.error(f"验证失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "系统错误", f"验证过程中发生意外错误: {str(e)}")
            return False

    def _get_widget_value(self, widget, field):
        """从控件获取值（从原get_field_value方法提取）"""
        if isinstance(widget, QDateEdit):
            return widget.date().toString("yyyy-MM-dd")
        elif isinstance(widget, QTextEdit):
            return widget.toPlainText().strip()
        else:  # QLineEdit等
            return widget.text().strip()

    def get_parent(self):
        """安全获取父窗口引用"""
        return self.parent_ref() if self.parent_ref else None
    
    def _on_save(self):
        """保存按钮处理"""
        if not self._validate_data():
            return
            
        try:
            data = self._validated_data
            parent = self.get_parent()
            
            if not parent:
                QMessageBox.critical(self, "错误", "无法获取父窗口引用")
                return
                
            if self.mode == 'add':
                success = add_personnel(parent.db_path, data)
            else:
                success = update_personnel(
                    parent.db_path, 
                    self.row_data['身份证号'], 
                    data
                )
                
            if success:
                self.accept()
            else:
                QMessageBox.warning(self, "警告", "保存失败，请检查数据是否冲突")
        except Exception as e:
            logger.error(f"保存失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "保存错误", f"数据库保存失败: {str(e)}")

# --------------------------- 新增统计对话框 ---------------------------
class StatsDialog(QDialog):
    def __init__(self, parent=None, db_path=None):
        super().__init__(parent)
        self.db_path = db_path
        self.setWindowTitle("数据统计")
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        
        # 统计字段选择
        self.field_combo = QComboBox()
        self.field_combo.addItems(["身份", "籍贯","一级单位", "二级单位", "学历", "出生年份"])
        
        # 统计类型
        self.type_combo = QComboBox()
        self.type_combo.addItems(["计数", "列表"])
        
        # 条件过滤
        self.filter_label = QLabel("筛选条件:")
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("例如: 一级单位='办公室'")
        
        # 按钮
        btn_run = QPushButton("执行统计")
        btn_run.clicked.connect(self.run_stats)
        
        # 结果显示
        self.result_area = QTextEdit()
        self.result_area.setReadOnly(True)
        
        layout.addWidget(QLabel("统计字段:"))
        layout.addWidget(self.field_combo)
        layout.addWidget(QLabel("统计类型:"))
        layout.addWidget(self.type_combo)
        layout.addWidget(self.filter_label)
        layout.addWidget(self.filter_input)
        layout.addWidget(btn_run)
        layout.addWidget(self.result_area)
        
        self.setLayout(layout)
        self.resize(500, 400)

    def run_stats(self):
        try:
            field = self.field_combo.currentText()
            stat_type = self.type_combo.currentText()
            condition = self.filter_input.text().strip()
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # 处理特殊字段
                if field == "出生年份":
                    field_expr = "strftime('%Y', 出生日期)"
                else:
                    field_expr = field
                    
                # 构建安全查询
                where_clause = ""
                params = []
                if condition:
                    if "=" in condition:
                        col, val = condition.split("=", 1)
                        where_clause = f"WHERE {col.strip()} = ?"
                        params.append(val.strip())
                
                if stat_type == "计数":
                    query = f"SELECT {field_expr}, COUNT(*) FROM personnel {where_clause} GROUP BY {field_expr}"
                    cursor.execute(query, params)
                    results = cursor.fetchall()
                    
                    output = f"{field}统计结果（总计{len(results)}类）:\n"
                    for name, count in results:
                        output += f"{name if name else '空值'}: {count}人\n"
                
                self.result_area.setText(output)
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"统计失败: {str(e)}")

class SmartSpinBox(QSpinBox):
    def __init__(self, min_value, max_value, suffix="", parent=None):
        super().__init__(parent)
        self.setRange(min_value, max_value)
        self.setValue(1)  # 默认值设为1
        self.setSuffix(suffix)
        self.setAlignment(Qt.AlignRight)

# --------------------------- 增强版简单套打模块 ---------------------------
class SimpleTemplateDialog(QDialog):
    template_loaded = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        logger.debug("简单套打对话框初始化开始")
        self.lbl_current = QLabel("当前使用模板：无")
        self.templates = {}
        self.current_template = ""
        self.mappings = {}
        self.current_data = {}
        self.temp_files = []
        self.template_combo = QComboBox()
        self.field_table = QTableWidget()
        
        # 确保模板目录存在
        template_dir = resource_path('templates')
        os.makedirs(template_dir, exist_ok=True)
        
        # 检查默认模板
        default_template = os.path.join(template_dir, "default_template.xlsx")
        if not os.path.exists(default_template):
            # 创建默认模板
            wb = Workbook()
            ws = wb.active
            ws.title = "默认模板"
            ws['A1'] = "姓名: {姓名}"
            ws['A2'] = "身份证号: {身份证号}"
            ws['A3'] = "单位: {一级单位}"
            wb.save(default_template)
            logger.debug("已创建默认模板文件")
        
        self.load_templates()
        self.setup_ui()  # 初始化UI
        self.temp_dir = os.path.join(tempfile.gettempdir(), "archive_system")
        os.makedirs(self.temp_dir, exist_ok=True)
        
        # 尝试加载默认模板
        if os.path.exists(default_template):
            self.current_template = default_template
            self.lbl_current.setText(f"当前使用模板：默认模板")
            logger.debug("已加载默认模板")
            
        self.update_field_table()  # 初始化时更新字段表
        logger.debug("简单套打对话框初始化完成")

    # ----------------- 事件处理函数 -----------------


    def enhanced_export(self):
        """导出文件"""
        try:
            temp_path = self.generate_filled_excel()
            save_path, _ = QFileDialog.getSaveFileName(
                self, "保存文件", 
                os.path.expanduser("~/Desktop/套打结果.xlsx"),
                "Excel文件 (*.xlsx)"
            )
            if save_path:
                shutil.copy(temp_path, save_path)
                QMessageBox.information(self, "成功", f"文件已保存到：{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出错误: {str(e)}")

    def get_personnel_data_from_db(self, ids):
        """直接从数据库获取人员数据（增强稳定性）"""
        try:
            if not ids:
                return []
                
            db_path = get_db_path()
            logger.debug(f"从数据库获取人员数据，ID列表: {ids}")
            
            with sqlite3.connect(db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                placeholders = ','.join(['?'] * len(ids))
                query = f"SELECT * FROM personnel WHERE 身份证号 IN ({placeholders})"
                cursor.execute(query, ids)
                results = [dict(row) for row in cursor.fetchall()]
                logger.debug(f"从数据库获取到 {len(results)} 条记录")
                return results
        except Exception as e:
            logger.error(f"数据库查询失败: {str(e)}")
            QMessageBox.critical(self, "数据库错误", f"查询失败: {str(e)}")
            return []

    def setup_ui(self):
        """初始化UI界面（增强稳定性）"""
        try:
            self.setWindowTitle("模板套打系统")
            
            # 使用主垂直布局，字段映射区域占据主要空间
            main_layout = QVBoxLayout(self)  # 关键修改：直接设置给对话框
            main_layout.setContentsMargins(10, 10, 10, 10)
            
            # ===== 字段映射表 - 主要区域 =====
            field_group = QGroupBox("字段映射")
            field_layout = QVBoxLayout(field_group)
            
            # 添加搜索框
            search_layout = QHBoxLayout()
            search_layout.addWidget(QLabel("搜索字段:"))
            self.search_field = QLineEdit()
            self.search_field.setPlaceholderText("输入关键词过滤...")
            self.search_field.textChanged.connect(self.filter_fields)
            search_layout.addWidget(self.search_field)
            field_layout.addLayout(search_layout)
            
            # 字段表格 - 占据主要空间
            self.field_table = QTableWidget()
            self.field_table.setColumnCount(3)
            self.field_table.setHorizontalHeaderLabels(["模板字段", "映射字段", "示例值"])
            self.field_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
            self.field_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            self.field_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
            field_layout.addWidget(self.field_table)
            
            main_layout.addWidget(field_group, 4)  # 4份空间
            
            # ===== 模板操作和打印设置 - 次要区域 =====
            bottom_container = QWidget()
            bottom_layout = QHBoxLayout(bottom_container)
            
            # 模板操作区
            template_group = QGroupBox("模板操作")
            template_layout = QVBoxLayout(template_group)
            self.btn_store = QPushButton("模板仓库")
            self.btn_create = QPushButton("制作新模板")
            template_layout.addWidget(self.btn_store)
            template_layout.addWidget(self.btn_create)
            bottom_layout.addWidget(template_group)
            
            # 当前模板标签
            current_template_box = QGroupBox("当前模板")
            current_layout = QVBoxLayout(current_template_box)
            self.lbl_current = QLabel("无")
            self.lbl_current.setWordWrap(True)
            self.lbl_current.setStyleSheet("font-size: 10pt; padding: 5px;")
            current_layout.addWidget(self.lbl_current)
            bottom_layout.addWidget(current_template_box)
            
            # 打印设置区
            print_settings = QGroupBox("打印设置")
            setting_layout = QVBoxLayout(print_settings)
            
            # 打印份数
            copies_layout = QHBoxLayout()
            copies_layout.addWidget(QLabel("份数:"))
            self.copies_spin = SmartSpinBox(1, 999, "")
            self.copies_spin.setFixedWidth(80)
            copies_layout.addWidget(self.copies_spin)
            setting_layout.addLayout(copies_layout)
            
            # 双面打印
            self.duplex_check = QCheckBox("双面打印")
            setting_layout.addWidget(self.duplex_check)
            
            # 功能按钮
            btn_layout = QHBoxLayout()
            self.btn_config = QPushButton("保存配置")
            self.btn_preview = QPushButton("预览")
            self.btn_print = QPushButton("打印")
            self.btn_export = QPushButton("导出")
            
            btn_layout.addWidget(self.btn_config)
            btn_layout.addWidget(self.btn_preview)
            btn_layout.addWidget(self.btn_print)
            btn_layout.addWidget(self.btn_export)
            setting_layout.addLayout(btn_layout)
            
            bottom_layout.addWidget(print_settings)
            
            main_layout.addWidget(bottom_container, 1)  # 1份空间
            
            # ========= 事件绑定 =========
            self.btn_preview.clicked.connect(self.enhanced_preview)
            self.btn_print.clicked.connect(self.enhanced_print)
            self.btn_export.clicked.connect(self.enhanced_export)
            self.btn_config.clicked.connect(self.save_current_config)
            self.template_combo.currentIndexChanged.connect(self.on_template_selected)
            self.btn_store.clicked.connect(self.show_template_store)
            self.btn_create.clicked.connect(self.create_new_template)
            
            self.setLayout(main_layout)
            self.resize(1000, 700)  # 增大窗口尺寸以适应新布局
            
            logger.debug("简单套打UI初始化完成")
        except Exception as e:
            logger.error(f"简单套打UI初始化失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "UI错误", f"界面初始化失败: {str(e)}")
    
    def filter_fields(self, text):
        """根据搜索文本过滤显示字段"""
        text = text.strip().lower()
        for row in range(self.field_table.rowCount()):
            item = self.field_table.item(row, 0)
            if item:
                field_name = item.text().lower()
                self.field_table.setRowHidden(row, text and text not in field_name)   

    def get_field_mappings(self):
        """从表格获取字段映射关系"""
        mappings = {}
        for row in range(self.field_table.rowCount()):
            field_name_item = self.field_table.item(row, 0)
            if not field_name_item:
                continue
            field_name = field_name_item.text().strip()
            combo = self.field_table.cellWidget(row, 1)
            if combo and combo.currentText():
                mappings[field_name] = combo.currentText()
        return mappings

    def load_templates(self):
        """加载模板到下拉框"""
        config_path = os.path.join(resource_path('templates'), 'template_config.json')
        self.template_combo.clear()
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    self.templates = json.load(f)
                    self.template_combo.addItems(self.templates.keys())
            except Exception as e:
                logger.error(f"加载模板配置失败: {str(e)}")

    def show_template_store(self):
        """显示模板仓库"""
        try:
            dialog = TemplateStoreDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                selected_template = dialog.get_selected_template()
                if selected_template:
                    self.load_template_config(selected_template)
        except Exception as e:
            logger.error(f"打开模板仓库失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"打开模板仓库失败: {str(e)}")

    def create_new_template(self):
        """增强版创建新模板方法（解决闪退问题）"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "选择模板文件", 
                os.path.expanduser("~/Desktop"),
                "Excel文件 (*.xlsx *.xls)"
            )
            if not file_path:
                return
                
            # 获取模板文件名
            template_filename = os.path.basename(file_path)
            template_name = os.path.splitext(template_filename)[0]
            
            # 获取模板目录（使用资源路径函数）
            template_dir = resource_path('templates')
            os.makedirs(template_dir, exist_ok=True)
            dest_path = os.path.join(template_dir, template_filename)
            
            # 检查文件是否已存在
            if os.path.exists(dest_path):
                reply = QMessageBox.question(
                    self, "覆盖确认",
                    f"模板文件 '{template_filename}' 已存在，是否覆盖？",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.No:
                    return
            
            # 复制文件到模板目录
            try:
                shutil.copy2(file_path, dest_path)
            except PermissionError:
                QMessageBox.critical(self, "错误", "文件被占用或无写入权限")
                return
            except Exception as e:
                logger.error(f"复制模板文件失败: {str(e)}")
                QMessageBox.critical(self, "错误", f"复制文件失败: {str(e)}")
                return
            
            # 设置当前模板
            self.current_template = dest_path
            self.lbl_current.setText(f"当前使用模板：{template_name}")
            
            # 初始化字段映射表
            self.update_field_table()
            # 尝试自动匹配字段
            self.auto_map_fields()
            
            # 保存配置（新增方法调用）
            self.save_template_config(template_name, dest_path)
            
            QMessageBox.information(self, "成功", "模板创建并保存成功！")
        except Exception as e:
            logger.error(f"创建模板失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "错误", f"创建模板失败: {str(e)}")

    # 新增方法：专门保存模板配置
    def save_template_config(self, template_name, template_path):
        """保存模板配置到JSON文件"""
        try:
            # 获取字段映射
            mappings = self.get_field_mappings()
            
            # 构建配置对象
            config = {
                "template_path": template_path,
                "mappings": mappings,
                "create_time": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
            
            # 配置文件路径
            config_path = os.path.join(resource_path('templates'), 'template_config.json')
            
            # 加载现有配置
            all_configs = {}
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    all_configs = json.load(f)
            
            # 更新配置
            all_configs[template_name] = config
            
            # 保存配置
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(all_configs, f, indent=2, ensure_ascii=False)
            
            # 刷新模板列表
            self.load_templates()

            # 保存为上次使用的模板
            settings = QSettings("MyCompany", "ArchiveManager")
            settings.setValue("lastTemplate", template_name)
            
            logger.info(f"模板配置已保存: {template_name}")
        except Exception as e:
            logger.error(f"保存模板配置失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"保存配置失败: {str(e)}")

    # ----------------- 修改后的关键方法 -----------------
    def update_field_table(self, mappings=None):
        """更新字段映射表（增强错误处理）"""
        try:
            self.field_table.setRowCount(0)
            
            if not self.current_template or not os.path.exists(self.current_template):
                QMessageBox.warning(self, "警告", "模板文件不存在或路径无效")
                return
                
            # 加载工作簿
            wb = load_workbook(self.current_template)
            ws = wb.active
            
            # 收集所有字段
            fields = set()
            pattern = re.compile(r'\{\s*([^{}]+?)\s*\}')
            
            for row in ws.iter_rows():
                for cell in row:
                    cell_value = cell.value
                    if not cell_value or not isinstance(cell_value, str):
                        continue
                        
                    # 处理合并单元格
                    if ws.merged_cells:
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                                cell_value = top_left_cell.value
                                break
                    
                    matches = pattern.findall(str(cell_value))
                    fields.update(matches)
            
            # 获取示例数据
            example_data = self.get_example_data()
            
            # 排序字段
            sorted_fields = sorted(fields, key=lambda x: ('\u4e00' <= x[0] <= '\u9fff', x))
            
            # 填充表格
            for idx, field in enumerate(sorted_fields):
                self.field_table.insertRow(idx)
                clean_field = field.strip()
                
                # 第一列：模板字段
                field_item = QTableWidgetItem(clean_field)
                self.field_table.setItem(idx, 0, field_item)
                
                # 第二列：带搜索的下拉框
                combo = QComboBox()
                combo.setEditable(True)
                completer = QCompleter(field_manager.get_field_names())
                combo.setCompleter(completer)
                combo.addItems([""] + field_manager.get_field_names())
                
                # 设置当前映射
                if mappings and clean_field in mappings:
                    combo.setCurrentText(mappings[clean_field])
                    
                self.field_table.setCellWidget(idx, 1, combo)
                
                # 第三列：示例值
                mapped_field = combo.currentText()
                example_value = example_data.get(mapped_field, "无对应数据") if mapped_field else "请选择映射字段"
                
                example_item = QTableWidgetItem(str(example_value))
                if "无" in example_value:
                    example_item.setForeground(QColor(255, 0, 0))
                self.field_table.setItem(idx, 2, example_item)
                
                # 连接信号以便示例值实时更新
                combo.currentTextChanged.connect(
                    lambda text, row=idx: self.update_example_value(row, text)
                )
            
            # 调整列宽
            self.field_table.setColumnWidth(0, 200)
            self.field_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            
        except Exception as e:
            logger.error(f"更新字段表失败: {str(e)}")
            QMessageBox.critical(
                self, "错误", 
                f"字段加载失败: {str(e)}\n"
                "可能原因:\n"
                "1. Excel文件损坏\n"
                "2. 文件被其他程序占用\n"
                "3. 模板格式不支持"
            )

    # 新增辅助方法
    def update_example_value(self, row, field_name):
        """更新示例值（当用户选择映射字段时）"""
        example_data = self.get_example_data()
        example_value = example_data.get(field_name, "无对应数据") if field_name else "请选择映射字段"
        
        example_item = QTableWidgetItem(str(example_value))
        if "无" in example_value:
            example_item.setForeground(QColor(255, 0, 0))
        
        self.field_table.setItem(row, 2, example_item)

    def get_example_data(self):
        """获取示例数据（增强健壮性）"""
        try:
            # 尝试获取人员数据
            main_window = self.parent()
            if hasattr(main_window, 'get_personnel_data'):
                personnel_data = main_window.get_personnel_data([])
                if personnel_data:
                    return personnel_data[0]
        except:
            pass
        
        # 默认示例数据
        return {
            "姓名": "张三",
            "身份证号": "110101199003077654",
            "参加工作时间": "2010-08-01",
            "工作经历": "2010-2015 某公司\n2015-至今 现单位",
            "学历": "硕士研究生",
            "一级单位": "办公室",
            "二级单位": "人事科"
        }

    def on_template_selected(self, index):
        """当用户选择模板时更新界面并记录选择"""
        if index >= 0:
            template_name = self.template_combo.currentText()
            if self.load_template_config(template_name):
                # 保存为上次使用的模板
                settings = QSettings("MyCompany", "ArchiveManager")
                settings.setValue("lastTemplate", template_name)
            
    def load_config(self):
        """加载模板配置并恢复上次使用的模板"""
        try:
            config_path = os.path.join(resource_path('templates'), 'template_config.json')
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    self.templates = json.load(f)
            
            # 尝试加载上次使用的模板
            settings = QSettings("MyCompany", "ArchiveManager")
            last_template = settings.value("lastTemplate", "")
            
            if last_template and last_template in self.templates:
                self.load_template_config(last_template)
            else:
                # 没有上次模板时，尝试加载默认模板
                default_path = os.path.join(resource_path('templates'), "default_template.xlsx")
                if os.path.exists(default_path):
                    self.current_template = default_path
                    self.lbl_current.setText("当前使用模板：默认模板")
                    self.update_field_table()
        
        except Exception as e:
            logger.error(f"加载模板配置失败: {str(e)}")
            # 回退到默认模板
            default_path = os.path.join(resource_path('templates'), "default_template.xlsx")
            if os.path.exists(default_path):
                self.current_template = default_path
                self.lbl_current.setText("当前使用模板：默认模板")
                self.update_field_table()

    def auto_map_fields(self):
        """自动匹配相似字段"""
        template_fields = self.get_template_fields()
        system_fields = field_manager.get_field_names()
        
        for tpl_field in template_fields:
            clean_tpl = re.sub(r"[{}$]", "", tpl_field).strip()
            
            # 精确匹配优先
            if clean_tpl in system_fields:
                self.mappings[tpl_field] = clean_tpl
                continue
                
            # 模糊匹配（包含关系）
            for sys_field in system_fields:
                if sys_field in clean_tpl or clean_tpl in sys_field:
                    self.mappings[tpl_field] = sys_field
                    break
                
        self.update_field_table(self.mappings)

    def get_template_fields(self):
        """从当前模板提取字段列表"""
        if not self.current_template:
            return []
        
        try:
            wb = load_workbook(self.current_template)
            ws = wb.active
            fields = set()
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        matches = re.findall(r'\{(.+?)\}', cell.value)
                        fields.update(matches)
            return list(fields)
        except Exception as e:
            logger.error(f"获取模板字段失败: {str(e)}")
            return []

    def save_current_config(self):
        """保存当前模板配置"""
        if not self.current_template:
            return
            
        try:
            # 获取模板名称（从文件路径）
            template_name = os.path.splitext(os.path.basename(self.current_template))[0]
            
            # 保存配置
            self.save_template_config(template_name, self.current_template)
        except Exception as e:
            logger.error(f"保存配置失败: {str(e)}")

    def get_valid_mappings(self):
        """获取有效的字段映射（移除空映射）"""
        mappings = {}
        for row in range(self.field_table.rowCount()):
            field_name_item = self.field_table.item(row, 0)
            if not field_name_item:
                continue
                
            field_name = field_name_item.text().strip()
            combo = self.field_table.cellWidget(row, 1)
            
            if combo and combo.currentText():
                mappings[field_name] = combo.currentText()
        
        return mappings

    def enhanced_preview(self):
        """增强预览功能"""
        self.setCursor(Qt.WaitCursor)
        temp_path = None
        try:
            temp_path = self.generate_filled_excel()
            if sys.platform == 'win32':
                os.startfile(temp_path)
            elif sys.platform == 'darwin':
                subprocess.run(['open', temp_path], check=True)
            else:
                subprocess.run(['xdg-open', temp_path], check=True)
        except Exception as e:
            QMessageBox.critical(
                self, "预览失败",
                f"无法打开预览文件:\n{str(e)}\n文件路径: {temp_path if temp_path else '未生成'}"
            )
            logger.error(f"预览失败: {traceback.format_exc()}")
        finally:
            self.setCursor(Qt.ArrowCursor)

    def enhanced_print(self):
        """增强打印功能"""
        self.setCursor(Qt.WaitCursor)
        temp_path = None
        try:
            temp_path = self.generate_filled_excel()
            if sys.platform == 'win32':
                os.startfile(temp_path, "print")
            else:
                QMessageBox.information(
                    self, "打印",
                    f"文件已生成，请手动打印:\n{temp_path}"
                )
        except Exception as e:
            QMessageBox.critical(
                self, "打印失败",
                f"打印过程中出错:\n{str(e)}\n文件路径: {temp_path if temp_path else '未生成'}"
            )
            logger.error(f"打印失败: {traceback.format_exc()}")
        finally:
            self.setCursor(Qt.ArrowCursor)

    def enhanced_export(self):
        """导出填充后的文件"""
        try:
            temp_path = self.generate_filled_excel()
            save_path, _ = QFileDialog.getSaveFileName(
                self, "保存文件", 
                os.path.expanduser("~/Desktop/套打结果.xlsx"),
                "Excel文件 (*.xlsx)"
            )
            if save_path:
                shutil.copy(temp_path, save_path)
                QMessageBox.information(self, "成功", f"文件已保存到：{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")

    def generate_filled_excel(self):
        """生成填充后的Excel文件（修复缩进问题和闪退问题）"""
        try:
            # 参数校验
            if not self.current_template or not os.path.exists(self.current_template):
                logger.error(f"模板文件不存在: {self.current_template}")
                raise ValueError("未选择模板文件或模板文件不存在")
                
            # 获取字段映射
            mappings = self.get_valid_mappings()
            logger.debug(f"字段映射: {json.dumps(mappings, indent=2)}")
            
            # 获取人员数据
            personnel_data = []
            main_window = self.parent()
            
            if main_window and hasattr(main_window, 'get_selected_personnel_ids'):
                selected_ids = main_window.get_selected_personnel_ids()
                
                # 确保调用正确的方法
                if main_window and hasattr(main_window, 'get_personnel_data'):
                    personnel_data = main_window.get_personnel_data(selected_ids)
                    logger.debug(f"从主窗口获取到 {len(personnel_data)} 条人员数据")
                else:
                    # 备用方案：直接查询数据库
                    logger.warning("主窗口没有get_personnel_data方法，使用备用方案")
                    personnel_data = self.get_personnel_data_from_db(selected_ids)
            else:
                logger.error("无法获取主窗口的选中人员ID方法")
                QMessageBox.warning(self, "数据错误", "无法获取人员数据，请确保已选择人员")
                return None
            
            # 如果没有获取到数据，使用示例数据
            if not personnel_data:
                logger.warning("未获取到人员数据，使用示例数据")
                QMessageBox.warning(self, "数据警告", "未获取到人员数据，使用示例数据预览")
                personnel_data = [self.get_example_data()]
            
            # 创建临时文件
            temp_filename = f"preview_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            temp_path = os.path.join(self.temp_dir, temp_filename)
            
            # 确保临时目录存在
            os.makedirs(self.temp_dir, exist_ok=True)
            
            # 复制模板到临时文件
            try:
                shutil.copy2(self.current_template, temp_path)
            except Exception as e:
                logger.error(f"复制模板失败: {str(e)}")
                QMessageBox.critical(self, "文件错误", f"无法创建临时文件: {str(e)}")
                return None
            
            # 填充数据（支持多行数据）
            try:
                wb = load_workbook(temp_path)
                ws = wb.active
                pattern = re.compile(r'\{\s*([^{}]+?)\s*\}')
                
                for row_idx, person in enumerate(personnel_data):
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                # 替换所有匹配的占位符 - 修复关键错误
                                new_value = pattern.sub(
                                    lambda m: str(person.get(mappings.get(m.group(1).strip(), ""), ""),
                                    cell.value
                                ))
                                if new_value != cell.value:  # 仅修改有变化的单元格
                                    ws.cell(row=cell.row, column=cell.column, value=new_value)
                
                wb.save(temp_path)
                wb.close()
                logger.info(f"成功生成Excel文件: {temp_path}")
                return temp_path
            except Exception as e:
                logger.error(f"填充Excel失败: {traceback.format_exc()}")
                QMessageBox.critical(self, "生成错误", f"填充Excel失败: {str(e)}")
                return None
                
        except Exception as e:
            logger.error(f"生成Excel失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "错误", f"生成Excel失败: {str(e)}")
            return None

class TemplateStoreDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("模板仓库管理")
        self.setup_ui()
        self.load_templates()
        
    def setup_ui(self):
        """初始化界面布局"""
        layout = QVBoxLayout()
        
        # 模板列表
        self.list_widget = QListWidget()
        self.list_widget.itemDoubleClicked.connect(self.use_selected_template)
        layout.addWidget(QLabel("双击模板名称使用："))
        layout.addWidget(self.list_widget)

        # 按钮区域
        btn_layout = QHBoxLayout()
        btn_del = QPushButton("删除模板")
        btn_del.clicked.connect(self.delete_template)
        btn_use = QPushButton("使用模板")
        btn_use.clicked.connect(self.use_selected_template)
        btn_refresh = QPushButton("刷新列表")
        btn_refresh.clicked.connect(self.load_templates)
        
        btn_layout.addWidget(btn_del)
        btn_layout.addWidget(btn_use)
        btn_layout.addWidget(btn_refresh)
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
        self.resize(500, 400)

    def load_templates(self):
        """加载所有模板配置"""
        self.list_widget.clear()
        config_path = os.path.join(resource_path('templates'), 'template_config.json')
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    self.templates = json.load(f)
                for template_name in self.templates.keys():
                    self.list_widget.addItem(template_name)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"加载模板配置失败: {str(e)}")

    def get_selected_template(self):
        """获取当前选中的模板名称"""
        selected = self.list_widget.currentItem()
        return selected.text() if selected else None

    def use_selected_template(self):
        """使用选中的模板（修复父窗口引用问题）"""
        selected = self.list_widget.currentItem()
        if not selected:
            return

        try:
            # 通过窗口层级获取主窗口
            parent_dialog = self.parent()  # SimpleTemplateDialog
            main_window = None
            current_parent = self.parent()
            while current_parent:
                if hasattr(current_parent, 'simple_template_dialog'):
                    main_window = current_parent
                    break
                current_parent = current_parent.parent()
            
            # 备用方案：遍历应用所有窗口
            if not main_window:
                for widget in QApplication.topLevelWidgets():
                    if hasattr(widget, 'simple_template_dialog'):
                        main_window = widget
                        break
            
            # 最终校验
            if not main_window or not hasattr(main_window, 'simple_template_dialog'):
                logger.error("主窗口组件丢失，当前对象树：\n%s", 
                        "\n".join(self.get_parent_chain()))
                raise RuntimeError("核心组件初始化失败，请重启应用")

            template_name = selected.text()
            template_dialog = main_window.simple_template_dialog
            success = template_dialog.load_template_config(template_name)
            
            if success:
                self.accept()
                QMessageBox.information(self, "成功", f"模板加载成功: {template_name}")
                # 刷新父对话框
                parent_dialog.update_field_table()
            else:
                QMessageBox.warning(self, "错误", "模板配置文件可能已损坏")
                
        except Exception as e:
            logger.error(f"模板加载失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "错误", f"加载失败: {str(e)}\n请检查：\n1. 模板文件是否存在\n2. 字段映射是否完整")

    def delete_template(self):
        """删除选中的模板（增强健壮性）"""
        selected = self.list_widget.currentItem()
        if not selected:
            return
            
        template_name = selected.text()
        confirm = QMessageBox.question(
            self, "确认删除", 
            f"确定删除模板 '{template_name}' 吗？此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if confirm == QMessageBox.Yes:
            try:
                config_path = os.path.join(resource_path('templates'), 'template_config.json')
                with open(config_path, 'r', encoding='utf-8') as f:
                    configs = json.load(f)
                
                if template_name in configs:
                    # 删除模板文件（添加异常处理）
                    template_path = configs[template_name]['template_path']
                    if os.path.exists(template_path):
                        try:
                            os.remove(template_path)
                        except PermissionError:
                            QMessageBox.critical(self, "错误", "文件被占用或无删除权限")
                            return
                    
                    # 删除配置
                    del configs[template_name]
                    
                    # 保存更新后的配置
                    with open(config_path, 'w', encoding='utf-8') as f:
                        json.dump(configs, f, indent=2, ensure_ascii=False)
                    
                    # 刷新列表并提示
                    self.load_templates()
                    QMessageBox.information(self, "成功", f"模板 '{template_name}' 已删除")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"删除失败: {str(e)}")

# --------------------------- 其他新增组件 ---------------------------
class SearchableComboBox(QComboBox):
    """支持搜索的下拉框（独立组件）"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.completer = QCompleter(self)
        self.completer.setCompletionMode(QCompleter.PopupCompletion)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.setCompleter(self.completer)
        
    def addItems(self, texts):
        super().addItems(texts)
        self.completer.setModel(self.model())

class TemplateTypeComboBox(QComboBox):
    """支持分类的模板选择框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.categories = {}
        self.load_config()
        
    def load_config(self):
        """加载模板分类配置（自动生成默认文件）"""
        config_path = resource_path('config/template_categories.json')
        default_config = {
            "干部档案": {
                "icon": ":/icons/cadre.png",
                "templates": [
                    {"name": "干部信息表", "config": "cadre_info.json"},
                    {"name": "履历表", "config": "resume.json"}
                ]
            }
        }
        
        try:
            if not os.path.exists(config_path):
                os.makedirs(os.path.dirname(config_path), exist_ok=True)
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(default_config, f, indent=2, ensure_ascii=False)
            
            with open(config_path, 'r', encoding='utf-8') as f:
                self.categories = json.load(f)
            self.refresh()
        except Exception as e:
            logger.error(f"加载模板配置失败: {str(e)}")
            self.categories = default_config  # 使用默认配置

    def refresh(self):
        """刷新模板列表"""
        self.clear()
        for cat, props in self.categories.items():
            self.addItem(QIcon(props['icon']), cat)
            for tpl in props['templates']:
                self.addItem(f"    {tpl['name']}", tpl['config'])

class FieldMappingTable(QTableWidget):
    """增强型字段映射表"""
    def __init__(self):
        super().__init__(0, 3)
        self.setHorizontalHeaderLabels(["模板字段", "映射字段", "示例值"])
        self.setDragDropMode(QAbstractItemView.DropOnly)
        self.setup_validation()

    def setup_validation(self):
        """设置字段验证规则"""
        delegate = FieldMappingDelegate()
        self.setItemDelegateForColumn(1, delegate)

    def dropEvent(self, event):
        """处理字段拖放"""
        source = event.source()
        if isinstance(source, FieldListWidget):
            item = source.currentItem()
            row = self.rowCount()
            self.insertRow(row)
            self.setItem(row, 0, QTableWidgetItem(item.text()))
            combo = QComboBox()
            combo.addItems(field_manager.get_field_names())
            self.setCellWidget(row, 1, combo)
            self.setItem(row, 2, QTableWidgetItem("示例值"))

class FieldMappingDelegate(QStyledItemDelegate):
    """字段映射验证委托"""
    def createEditor(self, parent, option, index):
        editor = QComboBox(parent)
        editor.addItems([""] + field_manager.get_field_names())
        return editor


# ======================= 修改 ArchiveManager 类 =======================
class ArchiveManager(QMainWindow):
    def __init__(self):
        super().__init__()
        # 初始化状态变量
        self.last_sorted = ""  # 记录上次排序的字段
        self.sort_asc = True   # 默认升序排序
        
        # 在方法开头添加字体设置
        app_font = QFont("Microsoft YaHei", 14)
        QApplication.setFont(app_font)
        
        try:
            # 确保全局field_manager已初始化
            global field_manager
            if field_manager is None:
                field_manager = FieldManager()
                
            # 在初始化UI之前初始化数据库
            init_database()
            self.db_path = get_db_path()
            
            # 先恢复状态（会调用init_ui）
            self.restore_state()  # ✅ 关键修改：先恢复状态再加载数据
            
            # 状态恢复中已经加载了数据
        except Exception as e:
            logger.error(f"初始化失败: {str(e)}")
            QMessageBox.critical(
                self, "初始化失败",
                f"程序初始化失败: {str(e)}\n\n"
                f"详细信息: {traceback.format_exc()}"
            )
            sys.exit(1)
            
        # 连接表头点击事件
        self.table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        # 确保在最开始设置窗口属性
        self.setAttribute(Qt.WA_DeleteOnClose)
        self.restore_state()  # ✅ 关键修改：先恢复状态再加载数据
        
    def init_chinese_font(self):
        """强制注册中文字体（从TemplateDesigner移动过来）"""
        try:
            # 尝试加载常见中文字体
            font_paths = [
                r"C:\Windows\Fonts\simsun.ttc",  # Windows 宋体
                r"C:\Windows\Fonts\simhei.ttf",   # Windows 黑体
                "/System/Library/Fonts/STHeiti Medium.ttc",  # Mac
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"  # Linux
            ]
            
            # 检查系统是否包含宋体字体
            available_fonts = QFontDatabase().families()
            if "宋体" not in available_fonts:
                QMessageBox.warning(
                    self, "字体缺失",
                    "系统缺少宋体字体，界面显示可能异常\n"
                    "建议安装中文字体包"
                )
        except Exception as e:
            logger.error(f"字体初始化失败: {str(e)}")

    def create_button(self, text, selector, handler):
        """创建并配置一个按钮"""
        button = QPushButton(text)
        button.setObjectName(selector)  # 设置对象名以便样式表匹配
        button.clicked.connect(handler)
        
        # +++ 移除原有的样式表设置，因为全局样式已经设置 +++
        # 仅设置最小尺寸
        button.setMinimumSize(120, 45)  # 增大按钮最小尺寸
        
        return button

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)
        
        # 标题栏
        title_layout = QHBoxLayout()
        title_label = QLabel("人事档案管理系统")
        # 增大标题字体
        title_font = QFont()
        title_font.setPointSize(45)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setStyleSheet("""
            QLabel {
                color: #3f51b5;
                padding:18px 0;
                font-size: 16pt;                  
            }
        """)
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        # 增大用户信息字体
        user_info = QLabel(f"管理员 | {datetime.now().strftime('%Y-%m-%d')}")
        user_info_font = QFont()
        user_info_font.setPointSize(12)
        user_info.setFont(user_info_font)
        user_info.setStyleSheet("color: #666;")
        title_layout.addWidget(user_info)
        
        main_layout.addLayout(title_layout)

        # 搜索栏
        search_group = QGroupBox("人员检索")
        # 增大组框标题字体
        search_group.setStyleSheet("QGroupBox { font-size: 18px; }")
        search_layout = QVBoxLayout(search_group)
        search_layout.setContentsMargins(15, 20, 15, 15)
        
        search_input_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入关键字（姓名、身份证号、单位，空格分隔）")
        # 设置搜索框本身的样式表
        self.search_input.setStyleSheet("""
            QLineEdit {
                font-size: 20px;
                padding: 15px;
                min-height: 40px;
            }
        """)
        
        search_btn = QPushButton("搜索")
        reset_btn = QPushButton("取消搜索")
        # ==== 新增代码：增大搜索框字体 ====
        search_font = QFont()
        search_font.setPointSize(25)  # 设置字号为18
        self.search_input.setFont(search_font)
        
        search_btn = QPushButton("搜索")
        search_btn.setObjectName("searchBtn")
        search_btn.clicked.connect(self.search_personnel)
        
        reset_btn = QPushButton("取消搜索")
        reset_btn.setObjectName("resetBtn")
        reset_btn.clicked.connect(self.reset_search)
        
        search_input_layout.addWidget(self.search_input)
        search_input_layout.addWidget(search_btn)
        search_input_layout.addWidget(reset_btn)
        
        search_layout.addLayout(search_input_layout)
        main_layout.addWidget(search_group)

        # 全选/取消全选按钮
        select_btn_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("全选")
        self.select_all_btn.setObjectName("selectAllBtn")
        self.select_all_btn.clicked.connect(self.select_all_rows)
        
        self.deselect_all_btn = QPushButton("取消全选")
        self.deselect_all_btn.setObjectName("deselectAllBtn")
        self.deselect_all_btn.clicked.connect(self.deselect_all_rows)
        
        select_btn_layout.addWidget(self.select_all_btn)
        select_btn_layout.addWidget(self.deselect_all_btn)
        select_btn_layout.addStretch()
        
        main_layout.addLayout(select_btn_layout)

        # 数据表格
        self.table = QTableWidget()
        self.table.setColumnCount(len(field_manager.fields) + 1)
        headers = ["选择"] + field_manager.get_field_names()
        self.table.setHorizontalHeaderLabels(headers)
        
        # 增大表格字体
        self.table.setStyleSheet("""
            QTableWidget {
                font-size: 18px;  /* 从14增大到18 */
            }
            QHeaderView::section {
                font-size: 18px;  /* 从14增大到18 */
                height: 40px;     /* 增加表头高度适应大字体 */
            }
        """)
        
        self.table.setColumnWidth(0, 50)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)  # 可手动调整
        self.table.horizontalHeader().setStretchLastSection(True)  # 最后一列自动拉伸
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.doubleClicked.connect(self.edit_row)
        
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
        main_layout.addWidget(self.table, 1)

        # 按钮工具栏
        btn_group = QGroupBox("操作")
        # 增大组框标题字体
        btn_group.setStyleSheet("QGroupBox { font-size: 14px; }")
        btn_layout = QHBoxLayout(btn_group)
        btn_layout.setContentsMargins(15, 15, 15, 15)
        
        # 使用新创建的create_button方法
        self.add_btn = self.create_button("新增", "#addBtn", self.open_add_dialog)
        self.delete_btn = self.create_button("删除", "#deleteBtn", self.delete_rows)
        self.import_btn = self.create_button("导入", "#importBtn", self.import_data)
        self.export_btn = self.create_button("导出", "#exportBtn", self.export_to_excel)
        self.stats_btn = self.create_button("统计", "#statsBtn", self.open_stats_dialog)
        self.simple_template_btn = self.create_button("简单套打", "#simplePrintBtn", self.open_simple_template)
        self.advanced_template_btn = self.create_button("高级套打", "#advancedPrintBtn", self.open_advanced_template)
        self.sort_btn = self.create_button("排序", "#sortBtn", self.open_sort_dialog)
        self.field_mgr_btn = self.create_button("字段管理", "#fieldMgrBtn", self.open_field_manager)
        self.backup_btn = self.create_button("备份", "#backupBtn", self.backup_database)
        
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.delete_btn)
        btn_layout.addWidget(self.import_btn)
        btn_layout.addWidget(self.export_btn)
        btn_layout.addWidget(self.stats_btn)
        btn_layout.addWidget(self.sort_btn)  # 新增排序按钮
        btn_layout.addWidget(self.simple_template_btn)
        btn_layout.addWidget(self.advanced_template_btn)
        btn_layout.addWidget(self.field_mgr_btn)
        btn_layout.addWidget(self.backup_btn)
        
        main_layout.addWidget(btn_group)
        self.last_sorted = ""  # 确保属性存在
        self.sort_asc = True

        self.load_data()
        
        self.status_bar = self.statusBar()
        self.status_bar.showMessage("就绪", 5000)

    


    def get_app_style(self):
        return """
            /* 全局样式 */
            QWidget {
                font-family: 'Microsoft YaHei', 'Segoe UI', sans-serif;
                font-size: 18px; /* 增大全局字体 */
                background-color: #f5f7fa;
                color: #333;
            }
            
            /* +++ 新增：全局按钮样式 +++ */
            QPushButton {
                font-size: 24px;  /* 增大按钮字体 */
                padding: 10px 20px;  /* 增大内边距 */
                min-height: 45px;   /* 设置最小高度 */
                min-width: 100px;   /* 设置最小宽度 */
                border-radius: 4px;
            }
            
            QDialog, QMessageBox {
                font-size: 14px; /* 增大对话框字体 */
            }
            
            QTableWidget {
                font-size: 14px; /* 增大表格字体 */
            }
            
            QLabel {
                font-size: 14px; /* 增大标签字体 */
            }
            
            QLineEdit, QComboBox, QTextEdit, QDateEdit {
                font-size: 18px; /* 增大输入控件字体 */
            }
            
            /* 彩色按钮 */
            #selectAllBtn {
                background-color: #4CAF50;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #deselectAllBtn {
                background-color: #f44336;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #addBtn {
                background-color: #2196F3;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #deleteBtn {
                background-color: #f44336;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #importBtn {
                background-color: #FF9800;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #exportBtn {
                background-color: #4CAF50;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #statsBtn {
                background-color: #9C27B0;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #simplePrintBtn {
                background-color: #009688;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #advancedPrintBtn {
                background-color: #795548;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #fieldMgrBtn {
                background-color: #607D8B;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #backupBtn {
                background-color: #FF5722;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #searchBtn {
                background-color: #03A9F4;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            #resetBtn {
                background-color: #9E9E9E;
                color: white;
                font-size: 20px; /* +++ 增大按钮字体 +++ */
            }
            
            QPushButton:hover {
                opacity: 0.9;
                border: 1px solid #a0a0a0;
            }
            
            QGroupBox {
                border: 1px solid #d1d5db;
                border-radius: 8px;
                margin-top: 15px;
                padding-top: 15px;
                font-weight: bold;
                background-color: white;
                font-size: 18px; /* 增大组框字体 */
            }
            
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
                background-color: white;
                font-size: 14px; /* 增大组框标题字体 */
            }
            
            QTableWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                gridline-color: #e0e0e0;
                alternate-background-color: #f9f9f9;
                font-size: 14px; /* 确保表格字体大小 */
            }
            
            QHeaderView::section {
                background-color: #5c6bc0;
                color: white;
                padding: 8px;
                border: none;
                font-size: 14px; /* 增大表头字体 */
                font-weight: 500;
            }
            
            QLineEdit, QComboBox {
                border: 1px solid #d1d5db;
                border-radius: 4px;
                padding: 8px;
                background-color: white;
                font-size: 14px; /* 增大输入控件字体 */
                min-height: 36px;
            }
            
            QLineEdit:focus, QComboBox:focus {
                border: 1px solid #64b5f6;
            }
            
            QLabel {
                font-size: 14px; /* 增大标签字体 */
                color: #444;
            }
        """

    # ======================= 状态保存与恢复 =======================
    def save_current_state(self):
            """保存当前应用程序状态（增强选中行保存）"""
            settings = QSettings("MyCompany", "ArchiveManager")
            
            # 保存搜索条件
            settings.setValue("searchText", self.search_input.text())
            
            # 保存排序状态
            settings.setValue("lastSorted", self.last_sorted)
            settings.setValue("sortAsc", self.sort_asc)
            
            # 保存选中的行（使用身份证号）
            selected_ids = self.get_selected_personnel_ids()
            settings.setValue("selectedIds", json.dumps(selected_ids))
            
            # 保存表格列宽
            column_widths = {}
            for col in range(self.table.columnCount()):
                column_widths[col] = self.table.columnWidth(col)
            settings.setValue("columnWidths", json.dumps(column_widths))
            
            # 保存表格水平滚动条位置
            settings.setValue("horizontalScroll", self.table.horizontalScrollBar().value())
            
            # 保存窗口几何状态
            settings.setValue("windowGeometry", self.saveGeometry())
            settings.setValue("windowState", self.saveState())
            
            # +++ 新增：保存模板设计器状态 +++
            if hasattr(self, 'simple_template_dialog') and self.simple_template_dialog:
                self.simple_template_dialog.save_current_config()
            if hasattr(self, 'advanced_template_dialog') and self.advanced_template_dialog:
                self.advanced_template_dialog.save_template_state()
                
            logger.info("当前状态已保存")
        
    def restore_state(self):
        """恢复上次保存的应用程序状态"""
        settings = QSettings("MyCompany", "ArchiveManager")
        
        # 恢复窗口几何状态（必须在UI初始化前）
        self.restoreGeometry(settings.value("windowGeometry", QByteArray()))
        self.restoreState(settings.value("windowState", QByteArray()))
        
        # 现在初始化UI
        self.init_ui()
        
        # 恢复数据相关状态
        search_text = settings.value("searchText", "")
        self.search_input.setText(search_text)
        
        self.last_sorted = settings.value("lastSorted", "")
        self.sort_asc = settings.value("sortAsc", True, type=bool)
        
        # 恢复列宽
        try:
            column_widths = json.loads(settings.value("columnWidths", "{}"))
            for col, width in column_widths.items():
                self.table.setColumnWidth(int(col), width)
        except:
            pass
        
        # 加载数据（使用恢复的排序状态）
        self.load_data(self.last_sorted if self.last_sorted else "")
        
        # 恢复水平滚动条位置
        scroll_pos = settings.value("horizontalScroll", 0, type=int)
        self.table.horizontalScrollBar().setValue(scroll_pos)
        
        # 恢复选中行
        QTimer.singleShot(100, self.restore_selection_state)
        
        # 恢复模板设计器状态
        if hasattr(self, 'simple_template_dialog') and self.simple_template_dialog:
            self.simple_template_dialog.load_config()
        if hasattr(self, 'advanced_template_dialog') and self.advanced_template_dialog:
            self.advanced_template_dialog.restore_template_state()
        
    def restore_selection_state(self):
        """恢复选中行状态（必须在数据加载后调用）"""
        if not hasattr(self, 'table') or self.table.rowCount() == 0:
            return
            
        try:
            settings = QSettings("MyCompany", "ArchiveManager")
            selected_ids = json.loads(settings.value("selectedIds", "[]"))
            
            if not selected_ids:
                return
                
            # 找到身份证号列的索引
            id_column = -1
            for col in range(self.table.columnCount()):
                header = self.table.horizontalHeaderItem(col).text()
                if header == "身份证号":
                    id_column = col
                    break
            
            if id_column == -1:
                return
                
            # 遍历表格行，恢复选中状态
            for row in range(self.table.rowCount()):
                item = self.table.item(row, id_column)
                if item and item.text() in selected_ids:
                    chk_widget = self.table.cellWidget(row, 0)
                    if chk_widget:
                        chk = chk_widget.findChild(QCheckBox)
                        if chk:
                            chk.setChecked(True)
        except Exception as e:
            logger.error(f"恢复选中状态失败: {str(e)}")


        
    # ======================= 数据加载与操作 =======================
    def load_data(self, order_by=""):
        """加载数据到表格，支持可选排序参数"""
        try:
            # 确保数据库连接正常
            if not os.path.exists(self.db_path):
                logger.warning("数据库文件不存在，尝试重新初始化")
                # +++ 修复：添加全局声明 +++
                global field_manager
                init_database(field_manager)  # ✅ 添加 field_manager 参数
                    
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                # 获取当前字段配置
                fields = field_manager.get_field_names()
                # 构建字段列表，使用反引号包裹字段名
                field_list = ', '.join([f'`{field}`' for field in fields])
                
                # 构建基础查询
                query = f"SELECT {field_list} FROM personnel"
                
                # 安全添加排序条件
                if order_by and isinstance(order_by, str):
                    # 验证并安全处理排序字段
                    valid_fields = []
                    for field in order_by.split(','):
                        field = field.strip()
                        # 处理DESC后缀
                        sort_direction = ""
                        if " DESC" in field:
                            field = field.replace(" DESC", "")
                            sort_direction = " DESC"
                        elif " ASC" in field:
                            field = field.replace(" ASC", "")
                            sort_direction = " ASC"
                        
                        if field in fields:
                            valid_fields.append(f"`{field}`{sort_direction}")
                    
                    if valid_fields:
                        query += " ORDER BY " + ", ".join(valid_fields)
                
                logger.debug(f"执行查询: {query}")
                cursor.execute(query)
                rows = cursor.fetchall()
                
            # 更新表格 - 关键修复
            self.table.setRowCount(0)
            self.table.setColumnCount(len(fields) + 1)  # +1 用于复选框列
            headers = ["选择"] + fields
            self.table.setHorizontalHeaderLabels(headers)
            
            # 设置复选框列宽度
            self.table.setColumnWidth(0, 50)
            
            # 填充数据
            for row_idx, row_data in enumerate(rows):
                self.table.insertRow(row_idx)
                
                # 添加复选框到第0列 - 关键修复
                chk_widget = QWidget()
                chk_layout = QHBoxLayout(chk_widget)
                chk_layout.setAlignment(Qt.AlignCenter)
                chk_layout.setContentsMargins(0, 0, 0, 0)
                checkbox = QCheckBox()
                chk_layout.addWidget(checkbox)
                self.table.setCellWidget(row_idx, 0, chk_widget)
                
                # 填充数据列 - 关键修复（从第1列开始）
                for col_idx, value in enumerate(row_data):
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row_idx, col_idx + 1, item)  # +1 跳过复选框列
            
            # 自动调整列宽
            self.table.resizeColumnsToContents()
            
            # 在数据加载完成后恢复选中状态
            self.restore_selection_state()
            
        except Exception as e:
            logger.error(f"加载数据失败: {traceback.format_exc()}")
            QMessageBox.critical(
                self, "错误", 
                f"加载数据失败: {str(e)}\n"
                "请检查数据库文件是否完整"
            )


    def get_personnel_data(self, ids=None):
        """根据身份证号列表获取人员数据"""
        try:
            if not ids or not isinstance(ids, list):
                return []
            
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                placeholders = ','.join(['?'] * len(ids))
                query = f"SELECT * FROM personnel WHERE 身份证号 IN ({placeholders})"
                cursor.execute(query, ids)
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"获取人员数据失败: {str(e)}")
            return []



    def select_all_rows(self):
        for row in range(self.table.rowCount()):
            chk_widget = self.table.cellWidget(row, 0)  # 改为第0列
            if chk_widget and chk_widget.findChild(QCheckBox):
                chk = chk_widget.findChild(QCheckBox)
                if chk:
                    chk.setChecked(True)

    def deselect_all_rows(self):
        for row in range(self.table.rowCount()):
            chk_widget = self.table.cellWidget(row, 0)  # 改为第0列
            if chk_widget and chk_widget.findChild(QCheckBox):
                chk = chk_widget.findChild(QCheckBox)
                if chk:
                    chk.setChecked(False)

    def show_context_menu(self, pos):
        """修复右键菜单的行号获取逻辑"""
        try:
            # 获取正确的行号 (使用鼠标位置而不是选中项)
            row = self.table.rowAt(pos.y())
            
            if row < 0:  # 如果没有点在行上，不显示菜单
                return
                
            menu = QMenu()
            
            
            # 只保留删除和导出功能
            delete_action = QAction("删除", self)
            export_action = QAction("导出选中", self)
            
            delete_action.triggered.connect(lambda checked, r=row: self.delete_selected_row(r))
            export_action.triggered.connect(self.export_selected)
            
            menu.addActions([delete_action, export_action])
            menu.exec_(self.table.mapToGlobal(pos))
        except Exception as e:
            logger.error(f"右键菜单错误: {str(e)}")

    def export_selected(self):
        """导出当前选中行"""
        self.export_to_excel()        

    def edit_selected_row(self, row):
        """修复编辑行逻辑 - 添加安全检查和简化逻辑"""
        try:
            # 安全检查：行号有效性
            if row < 0 or row >= self.table.rowCount():
                QMessageBox.warning(self, "警告", "请选择有效的行！")
                return
                
            # 收集行数据 - 简化逻辑
            row_data = {}
            field_names = field_manager.get_field_names()
            
            for col_idx, field_name in enumerate(field_names):
                # 注意：第0列是复选框，所以数据列从1开始
                item = self.table.item(row, col_idx + 1)
                row_data[field_name] = item.text() if item else ""
                
            # 创建编辑对话框 - 使用安全方式
            dialog = DynamicFormDialog(self, mode='edit', row_data=row_data)
            dialog.setAttribute(Qt.WA_DeleteOnClose)  # 确保关闭时释放资源
            
            # 安全显示对话框
            if dialog.exec_() == QDialog.Accepted:
                # 使用延迟刷新避免竞争条件
                QTimer.singleShot(100, self.load_data)
        except Exception as e:
            logger.error(f"编辑行数据失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"打开编辑界面失败: {str(e)}")
    
 
    def delete_selected_row(self, row):
        try:
            if row < 0 or row >= self.table.rowCount():
                QMessageBox.warning(self, "警告", "请选择有效的行！")
                return
                
            id_column = -1
            for col in range(self.table.columnCount()):
                header = self.table.horizontalHeaderItem(col).text()
                if header == "身份证号":
                    id_column = col
                    break
                    
            if id_column == -1:
                QMessageBox.warning(self, "错误", "未找到身份证号列")
                return
                
            item = self.table.item(row, id_column)
            if not item or not item.text().strip():
                QMessageBox.warning(self, "错误", "未找到有效的身份证号")
                return
                
            id_number = item.text().strip()
            
            confirm = QMessageBox.question(
                self, "确认删除", 
                f"确定删除这条记录吗？",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if confirm == QMessageBox.Yes:
                if delete_personnel(self.db_path, [id_number]):
                    self.table.removeRow(row)
                    QMessageBox.information(self, "成功", "删除成功！")
                else:
                    QMessageBox.warning(self, "错误", "数据库删除失败")
        except Exception as e:
            logger.error(f"删除记录失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"删除操作失败: {str(e)}")
    
    # ... 其他方法保持不变 ...
    # 然后添加这个方法：
    # 修改 ArchiveManager 类的 open_advanced_template 方法
    def open_advanced_template(self):
        try:
            selected_ids = self.get_selected_personnel_ids()
            if not selected_ids:
                QMessageBox.warning(self, "警告", "请先选择人员")
                return

            if not hasattr(self, 'advanced_template_dialog') or self.advanced_template_dialog is None:
                self.advanced_template_dialog = TemplateDesigner(
                    self, 
                    db_path=self.db_path,
                    selected_ids=selected_ids
                )
            else:
                # 更新选中的身份证号
                self.advanced_template_dialog.selected_ids = selected_ids
                # 重新加载人员数据
                self.advanced_template_dialog.load_personnel_data()

            self.advanced_template_dialog.show()
        except Exception as e:
            logger.error(f"打开高级套打失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"高级套打功能不可用: {str(e)}")

    def delete_rows(self):
        """删除选中的行"""
        try:
            # 动态查找身份证号列
            id_column = -1
            for col in range(self.table.columnCount()):
                header = self.table.horizontalHeaderItem(col).text()
                if header == "身份证号":
                    id_column = col
                    break
            if id_column == -1:
                raise ValueError("未找到身份证号列")

            # 获取选中行索引
            selected_rows = []
            for row in range(self.table.rowCount()):
                chk_widget = self.table.cellWidget(row, 0)
                if chk_widget and chk_widget.findChild(QCheckBox).isChecked():
                    selected_rows.append(row)
            
            if not selected_rows:
                QMessageBox.warning(self, "警告", "请先勾选要删除的行！")
                return

            # 收集身份证号
            ids = []
            for row in selected_rows:
                item = self.table.item(row, id_column)
                if item and item.text().strip():
                    ids.append(item.text().strip())

            if not ids:
                QMessageBox.warning(self, "警告", "未找到有效的身份证号")
                return

            # 确认对话框
            confirm = QMessageBox.question(
                self, "确认删除", 
                f"确定删除这 {len(ids)} 条记录吗？",
                QMessageBox.Yes | QMessageBox.No
            )

            if confirm == QMessageBox.Yes:
                if delete_personnel(self.db_path, ids):
                    # 倒序删除表格行
                    for row in sorted(selected_rows, reverse=True):
                        self.table.removeRow(row)
                    QMessageBox.information(self, "成功", "删除成功！")
                else:
                    QMessageBox.warning(self, "错误", "数据库删除失败")
        except Exception as e:
            logger.error(f"删除记录失败: {str(e)}")
            QMessageBox.critical(
                self, "错误", 
                f"删除操作失败:\n{str(e)}\n" 
                "可能原因:\n"
                "1. 数据库被其他程序锁定\n"
                "2. 字段配置不匹配"
            )

    def open_field_manager(self):
        try:
            dialog = FieldManagerDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                QMessageBox.information(
                    self, "提示", 
                    "字段配置已保存！\n程序将退出，请重新启动以应用变更。"
                )
                sys.exit(0)  # 强制退出，确保数据库重建
        except Exception as e:
            logger.error(f"打开字段管理失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"配置错误: {str(e)}")

    def open_simple_template(self):
        """打开简单套打对话框（增强版）"""
        try:
            # 直接初始化对话框
            if not hasattr(self, 'simple_template_dialog') or self.simple_template_dialog is None:
                self.simple_template_dialog = SimpleTemplateDialog(self)
                
            # 刷新人员选择
            selected_ids = self.get_selected_personnel_ids()
            if not selected_ids:
                reply = QMessageBox.question(
                    self, '确认', 
                    '未选择任何人员，将处理全部记录。是否继续？',
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.No:
                    return
            
            # 显示对话框
            self.simple_template_dialog.show()
            self.simple_template_dialog.activateWindow()
        except Exception as e:
            logger.error(f"打开简单套打失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "错误", f"功能不可用: {str(e)}")

    def search_personnel(self):
        """增强版搜索功能，支持组合条件"""
        keyword = self.search_input.text().strip()
        if not keyword:
            self.load_data()
            return
            
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # 构建动态查询条件
                conditions = []
                params = []
                
                # 拆分搜索词
                search_terms = keyword.split()
                
                for term in search_terms:
                    # 检查是否是身份证号部分（数字）
                    if term.isdigit():
                        conditions.append("身份证号 LIKE ?")
                        params.append(f"%{term}%")
                    else:
                        # 可能是姓名或单位
                        conditions.append("(姓名 LIKE ? OR 一级单位 LIKE ? OR 二级单位 LIKE ?)")
                        params.extend([f"%{term}%", f"%{term}%", f"%{term}%"])
                
                where_clause = " AND ".join(conditions)
                query = f'''
                    SELECT 档案编号, 姓名, 身份证号, 身份, 籍贯,一级单位, 二级单位, 
                    出生日期, 参加工作时间, 入党日期, 工作经历, 学历, 
                    档案流转记录, 电子档案, 备注 
                    FROM personnel 
                    WHERE {where_clause}
                '''
                
                cursor.execute(query, params)
                rows = cursor.fetchall()
                
            self.table.setRowCount(0)
            for row_data in rows:
                row = self.table.rowCount()
                self.table.insertRow(row)
                for col, value in enumerate(row_data):
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row, col, item)
        except Exception as e:
            logger.error(f"搜索失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"搜索失败: {str(e)}")

    def reset_search(self):
        self.search_input.clear()
        self.load_data()  # 使用默认排序
        QMessageBox.information(self, "提示", "已显示全部数据")      
    
    def import_data(self):
        """增强版导入方法"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "选择Excel文件", "", 
                "Excel文件 (*.xlsx *.xls)"
            )
            if not file_path:
                return
                
            # 创建进度对话框
            progress = QProgressDialog("导入数据中...", "取消", 0, 100, self)
            progress.setWindowTitle("导入进度")
            progress.setWindowModality(Qt.WindowModal)
            progress.setValue(10)
            QApplication.processEvents()
            
            # 执行导入
            count = import_from_excel(self.db_path, file_path)
            
            progress.setValue(100)
            progress.close()
            
            if count > 0:
                QMessageBox.information(
                    self, "成功", 
                    f"成功导入 {count} 条记录\n"
                    f"重复记录已自动跳过"
                )
                self.load_data()  # 刷新表格显示
            else:
                # 使用警告而不是错误，用户可以关闭
                QMessageBox.warning(
                    self, "警告", 
                    "没有导入任何记录\n"
                    "可能原因:\n"
                    "1. Excel格式不正确\n"
                    "2. 必填字段缺失\n"
                    "3. 数据冲突"
                )
        except Exception as e:
            # 使用警告而不是错误，用户可以关闭
            QMessageBox.warning(
                self, "导入问题",
                f"导入过程中遇到问题: {str(e)}\n"
                "部分记录可能未导入，请检查Excel文件格式"
            )
            logger.error(f"导入失败: {traceback.format_exc()}")

    def export_to_excel(self):
        """增强版导出方法，支持导出选中行"""
        try:
            # 获取选中行的身份证号
            selected_ids = self.get_selected_personnel_ids()
            
            # 构建查询条件
            condition = ""
            params = []
            if selected_ids:
                placeholders = ','.join(['?'] * len(selected_ids))
                condition = f" WHERE 身份证号 IN ({placeholders})"
                params = selected_ids
                
            # 获取所有字段名
            fields = field_manager.get_field_names()
            field_list = ', '.join([f'"{field}"' for field in fields])
            
            # 构建查询
            query = f"SELECT {field_list} FROM personnel{condition}"
            
            # 创建临时文件路径
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            default_name = f"人事档案导出_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            save_path, _ = QFileDialog.getSaveFileName(
                self, "保存文件", 
                os.path.join(desktop, default_name),
                "Excel文件 (*.xlsx)"
            )
            if not save_path:
                return
                
            # 添加进度提示
            progress = QProgressDialog("导出数据中...", "取消", 0, 100, self)
            progress.setWindowTitle("导出进度")
            progress.setWindowModality(Qt.WindowModal)
            progress.setValue(30)
            QApplication.processEvents()
            
            # 执行导出
            with sqlite3.connect(self.db_path) as conn:
                df = pd.read_sql_query(query, conn, params=params)
                df.to_excel(save_path, index=False, engine="openpyxl")
                
            progress.setValue(100)
            progress.close()
            
            QMessageBox.information(
                self, "成功", 
                f"文件已保存到：\n{save_path}\n"
                f"共导出 {len(df)} 条记录"
            )
            return True
            
        except Exception as e:
            logger.error(f"导出失败: {traceback.format_exc()}")
            QMessageBox.critical(
                self, "错误", 
                f"导出失败: {str(e)}\n"
                "建议操作:\n"
                "1. 关闭已打开的Excel文件\n"
                "2. 检查磁盘空间"
            )
            return False

    # 在 ArchiveManager 类中添加以下方法

    def save_current_state(self):
        """保存当前应用程序状态（增强选中行保存）"""
        settings = QSettings("MyCompany", "ArchiveManager")
        # 保存搜索条件
        settings.setValue("searchText", self.search_input.text())
        # 保存排序状态
        settings.setValue("lastSorted", self.last_sorted)
        settings.setValue("sortAsc", self.sort_asc)
        # 保存选中的行（使用身份证号）
        selected_ids = self.get_selected_personnel_ids()
        settings.setValue("selectedIds", json.dumps(selected_ids))
        # 保存表格列宽
        column_widths = {}
        for col in range(self.table.columnCount()):
            column_widths[col] = self.table.columnWidth(col)
        settings.setValue("columnWidths", json.dumps(column_widths))
        # 保存表格水平滚动条位置
        settings.setValue("horizontalScroll", self.table.horizontalScrollBar().value())
        # 保存窗口几何状态
        settings.setValue("windowGeometry", self.saveGeometry())
        settings.setValue("windowState", self.saveState())
        
        # +++ 新增：保存模板设计器状态 +++
        if hasattr(self, 'advanced_template_dialog') and self.advanced_template_dialog:
            self.advanced_template_dialog.save_template_state()

    def closeEvent(self, event):
        """关闭事件处理 - 添加保存提示"""
        reply = QMessageBox.question(
            self, "确认退出",
            "是否保存当前数据状态？\n保存后下次打开将恢复当前页面。",
            QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
            QMessageBox.Save
        )
        if reply == QMessageBox.Save:
            self.save_current_state()  # 保存当前状态
            event.accept()
        elif reply == QMessageBox.Discard:
            event.accept()
        else:
            event.ignore()

    

    def on_header_clicked(self, column):
        if column == 0:  # 跳过复选框列
            return
        
        try:
            field_name = self.table.horizontalHeaderItem(column).text()
            
            # 验证字段名有效性
            if field_name not in field_manager.get_field_names():
                logger.warning(f"无效的排序字段: {field_name}")
                return
                
            # 切换升序/降序
            if self.last_sorted == field_name:
                self.sort_asc = not self.sort_asc
            else:
                self.sort_asc = True
            
            # 构建安全的排序参数 - 直接传递字段名
            order = field_name  # 只传递字段名，不加引号
            if not self.sort_asc:
                order += " DESC"
            
            # 加载数据
            self.load_data(order)
            self.last_sorted = field_name
            
        except Exception as e:
            logger.error(f"排序失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"排序操作失败: {str(e)}")

    def get_selected_personnel_ids(self):
        """获取选中人员的身份证号列表"""
        ids = []
        # 找到身份证号列的索引
        id_column = -1
        for col in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(col).text()
            if header == "身份证号":
                id_column = col
                break
        
        if id_column == -1:
            return []
        
        # 收集选中行的身份证号
        for row in range(self.table.rowCount()):
            if self.table.cellWidget(row, 0).findChild(QCheckBox).isChecked():
                item = self.table.item(row, id_column)
                if item and item.text().strip():
                    ids.append(item.text().strip())
        
        return ids
    

    def open_add_dialog(self):
        """打开新增档案对话框"""
        try:
            dialog = DynamicFormDialog(self, mode='add')
            if dialog.exec_() == QDialog.Accepted:
                self.load_data()
        except Exception as e:
            logger.error(f"打开新增对话框失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"无法打开新增界面: {str(e)}")
    
 

    def open_import_dialog(self):
        dialog = ImportDialog(self, db_path=self.db_path)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()

    def open_stats_dialog(self):
        """打开统计对话框"""
        try:
            dialog = StatsDialog(self, db_path=self.db_path)
            dialog.exec_()
        except Exception as e:
            logger.error(f"打开统计对话框失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"无法打开统计界面: {str(e)}")

    def quick_print(self, template_name, ids):
        """快速打印已配置的模板"""
        try:
            # 加载模板配置
            template_dir = os.path.join(os.path.expanduser("~/Documents/人事档案系统"), "templates")
            config_path = os.path.join(template_dir, "template_config.json")
            
            with open(config_path, 'r', encoding='utf-8') as f:
                templates = json.load(f)
                
            if template_name not in templates:
                QMessageBox.warning(self, "警告", "模板不存在！")
                return
                
            config = templates[template_name]
            
            # 获取人员数据
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                placeholders = ','.join(['?'] * len(ids))
                cursor.execute(
                    f"SELECT * FROM personnel WHERE 身份证号 IN ({placeholders})", 
                    ids
                )
                data_list = [dict(row) for row in cursor.fetchall()]
                
            if not data_list:
                QMessageBox.warning(self, "警告", "没有可打印的数据！")
                return
                
            # 使用openpyxl填充并打印
            for data in data_list:
                wb = load_workbook(config['path'])
                ws = wb.active
                
                # 填充数据
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("{") and cell.value.endswith("}"):
                            field_name = cell.value[1:-1].strip()
                            if field_name in config['mappings']:
                                cell.value = data.get(config['mappings'][field_name], "")
            
                # 创建临时文件并打印
                temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                temp_path = temp_file.name
                temp_file.close()
                
                wb.save(temp_path)
                os.startfile(temp_path, "print")  # Windows系统
                
            QMessageBox.information(self, "成功", f"已发送{len(data_list)}条打印任务！")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"快速打印失败: {str(e)}")  

    def edit_selected_row(self):
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "警告", "请选择要编辑的行！")
            return
        self.edit_row()

    def edit_row(self):
        try:
            selected_items = self.table.selectedItems()
            if not selected_items:
                return
                
            row = selected_items[0].row()
            row_data = {}
            
            for col_idx, field in enumerate(field_manager.fields):
                item = self.table.item(row, col_idx + 1)  # +1跳过复选框列
                row_data[field['name']] = item.text() if item else ""
                
            dialog = DynamicFormDialog(self, mode='edit', row_data=row_data)  # 确保self作为parent传入
            if dialog.exec_() == QDialog.Accepted:
                self.load_data()
        except Exception as e:
            logger.error(f"编辑行数据失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"打开编辑界面失败: {str(e)}")
        
    def backup_database(self):
        backup_dir = os.path.join(os.path.expanduser("~/Documents/人事档案系统"), "backups")
        os.makedirs(backup_dir, exist_ok=True)  # ✅ 确保目录存在
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"archive_backup_{timestamp}.db")
            shutil.copy2(self.db_path, backup_path)
            QMessageBox.information(self, "成功", f"数据库已备份到：\n{backup_path}")
        except PermissionError:
            QMessageBox.critical(self, "权限错误", "无法写入备份目录，请以管理员身份运行程序")
        except Exception as e:
            logger.error(f"备份失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"备份失败：{str(e)}")   
    
    # 修改 open_sort_dialog 方法
    def open_sort_dialog(self):
        """打开排序对话框"""
        try:
            dialog = SortDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                order_by = dialog.get_sort_order()
                if order_by:
                    self.load_data(order_by)  # 现在可以正确传递参数了
        except Exception as e:
            logger.error(f"排序失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"排序操作失败: {str(e)}")
    
    def closeEvent(self, event):
        # 保存当前状态
        settings = QSettings("MyCompany", "ArchiveManager")
        settings.setValue("windowGeometry", self.saveGeometry())
        settings.setValue("windowState", self.saveState())
        event.accept()
        
    def showEvent(self, event):
        """显示事件 - 恢复上次状态"""
        super().showEvent(event)
        self.restore_state()
        self.restore_data_state()  # 新增：恢复数据状态

    def restore_data_state(self):
        """恢复应用程序数据状态（选中的行、列宽、滚动条位置等）"""
        settings = QSettings("MyCompany", "ArchiveManager")
        
        # 恢复搜索条件
        search_text = settings.value("searchText", "")
        self.search_input.setText(search_text)
        
        # 恢复排序状态
        self.last_sorted = settings.value("lastSorted", "")
        self.sort_asc = settings.value("sortAsc", True, type=bool)
        if self.last_sorted:
            self.load_data(self.last_sorted)
        
        # 恢复选中行
        QTimer.singleShot(100, self.restore_selection_state)
        
        # 恢复选中的行 - 关键修复
        try:
            selected_ids = json.loads(settings.value("selectedIds", "[]"))
            if selected_ids:
                # 找到身份证号列的索引
                id_column = -1
                for col in range(self.table.columnCount()):
                    header = self.table.horizontalHeaderItem(col).text()
                    if header == "身份证号":
                        id_column = col
                        break
                
                if id_column != -1:
                    for row in range(self.table.rowCount()):
                        item = self.table.item(row, id_column)
                        if item and item.text() in selected_ids:
                            chk_widget = self.table.cellWidget(row, 0)
                            if chk_widget:
                                chk = chk_widget.findChild(QCheckBox)
                                if chk:
                                    chk.setChecked(True)
        except Exception as e:
            logger.error(f"恢复选中状态失败: {str(e)}")
        
        # 恢复列宽
        try:
            column_widths = json.loads(settings.value("columnWidths", "{}"))
            for col, width in column_widths.items():
                self.table.setColumnWidth(int(col), width)
        except:
            pass
        
        # 恢复水平滚动条位置
        scroll_pos = settings.value("horizontalScroll", 0, type=int)
        self.table.horizontalScrollBar().setValue(scroll_pos)
        
        # 恢复排序状态（如果有）
        if self.last_sorted:
            self.load_data(self.last_sorted)
            
        # 如果有搜索条件，执行搜索
        if search_text:
            QTimer.singleShot(100, self.search_personnel)
            
        # +++ 新增：恢复模板设计器状态 +++
        if hasattr(self, 'simple_template_dialog') and self.simple_template_dialog:
            self.simple_template_dialog.load_config()
        if hasattr(self, 'advanced_template_dialog') and self.advanced_template_dialog:
            self.advanced_template_dialog.restore_template_state()

    def save_current_state(self):
        """保存当前应用程序状态（增强选中行保存）"""
        settings = QSettings("MyCompany", "ArchiveManager")
        
        # 保存搜索条件
        settings.setValue("searchText", self.search_input.text())
        
        # 保存排序状态
        settings.setValue("lastSorted", self.last_sorted)
        settings.setValue("sortAsc", self.sort_asc)
        
        # 保存选中的行（使用身份证号）
        selected_ids = []
        id_column = -1
        # 找到身份证号列
        for col in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(col).text()
            if header == "身份证号":
                id_column = col
                break
                
        if id_column != -1:
            for row in range(self.table.rowCount()):
                chk_widget = self.table.cellWidget(row, 0)
                if chk_widget:
                    chk = chk_widget.findChild(QCheckBox)
                    if chk and chk.isChecked():
                        item = self.table.item(row, id_column)
                        if item and item.text().strip():
                            selected_ids.append(item.text().strip())
        
        settings.setValue("selectedIds", json.dumps(selected_ids))
        
        # 保存表格列宽
        column_widths = {}
        for col in range(self.table.columnCount()):
            column_widths[col] = self.table.columnWidth(col)
        settings.setValue("columnWidths", json.dumps(column_widths))
        
        # 保存表格水平滚动条位置
        settings.setValue("horizontalScroll", self.table.horizontalScrollBar().value())
        
        # 保存窗口几何状态
        settings.setValue("windowGeometry", self.saveGeometry())
        settings.setValue("windowState", self.saveState())    
                 
    def closeEvent(self, event):
        """关闭事件处理 - 添加保存提示"""
        # 询问用户是否保存状态
        reply = QMessageBox.question(
            self, "确认退出",
            "是否保存当前数据状态？\n保存后下次打开将恢复当前页面。",
            QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
            QMessageBox.Save
        )
        
        if reply == QMessageBox.Save:
            # 保存当前状态
            self.save_current_state()
            event.accept()
        elif reply == QMessageBox.Discard:
            event.accept()
        else:  # Cancel
            event.ignore()

class SortDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("排序设置")
        # 获取所有字段名
        self.fields = field_manager.get_field_names()
        self.setup_ui()

    def setup_ui(self):
        layout = QFormLayout()
        
        # 四个排序字段选择
        self.primary_combo = QComboBox()
        self.primary_combo.addItems(["无"] + self.fields)  # 添加"无"选项
        self.secondary_combo = QComboBox()
        self.secondary_combo.addItems(["无"] + self.fields)
        self.tertiary_combo = QComboBox()
        self.tertiary_combo.addItems(["无"] + self.fields)
        self.quaternary_combo = QComboBox()
        self.quaternary_combo.addItems(["无"] + self.fields)
        
        layout.addRow("主要排序字段:", self.primary_combo)
        layout.addRow("次要排序字段:", self.secondary_combo)
        layout.addRow("第三排序字段:", self.tertiary_combo)
        layout.addRow("第四排序字段:", self.quaternary_combo)
        
        # 确认按钮
        btn_layout = QHBoxLayout()
        btn_confirm = QPushButton("确认")
        btn_confirm.clicked.connect(self.accept)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addWidget(btn_confirm)
        btn_layout.addWidget(btn_cancel)
        layout.addRow(btn_layout)
        
        self.setLayout(layout)
        self.resize(400, 300)
        
    def get_sort_order(self):
        orders = []
        for combo in [self.primary_combo, self.secondary_combo, 
                    self.tertiary_combo, self.quaternary_combo]:
            field = combo.currentText()
            if field != "无":
                orders.append(field)  # 只返回字段名，不加引号
                
        if orders:
            return ",".join(orders)  # 返回逗号分隔的字段列表
        return ""
# ---------------------------
class MovableFieldWidget(QWidget):
    def __init__(self, parent, field_name):
        super().__init__(parent)
        self.template_designer = parent
        self.field_name = field_name
        self.field_data = {
            'text': field_name,
            'x': 0, 'y': 0,
            'font': 'Arial',  # 默认字体
            'size': 10,
            'color': '#000000',
            'width': 120,
            'height': 40
        }
        self.init_ui(field_name)
        self.setCursor(Qt.SizeAllCursor)
        self.setAcceptDrops(True)
        self.dragging = False
        self.offset = QPoint()
        
        # 设置Qt端显示字体
        self.label.setFont(QFont("Arial", 10))
        self.label.setStyleSheet("color: black;")
        
        # 右键菜单
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_field_context_menu)

    def init_ui(self, text):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        
        self.label = QLabel(text)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("font-family: Arial; font-size: 10pt; color: #000000;")
        
        self.delete_btn = QPushButton("×")
        self.delete_btn.setFixedSize(20, 20)
        self.delete_btn.clicked.connect(self.safe_delete)
        
        layout.addWidget(self.label)
        layout.addWidget(self.delete_btn)
        
        self.setStyleSheet("""
            background: #e3f2fd;
            border: 1px solid #90caf9;
            border-radius: 3px;
            padding: 3px;
        """)

    def safe_delete(self):
        """安全删除方法"""
        try:
            self.deleteLater()
            if self.template_designer:
                self.template_designer.drop_area.update()
        except Exception as e:
            logger.error(f"删除字段失败: {str(e)}")

    def update_field_data(self):
        """实时更新字段坐标数据"""
        self.field_data.update({
            'x': self.x(),
            'y': self.y(),
            'width': self.width(),
            'height': self.height()
        })

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.dragging = True
            self.offset = event.pos()
            self.raise_()
            
    def mouseMoveEvent(self, event):
        if self.dragging and event.buttons() & Qt.LeftButton:
            new_pos = self.mapToParent(event.pos() - self.offset)
            self.move(new_pos)
            self.field_data.update({
                'x': new_pos.x(),
                'y': new_pos.y()
            })
            
    def mouseReleaseEvent(self, event):
        self.dragging = False

    def show_field_context_menu(self, pos):
        menu = QMenu(self)
        
        font_action = QAction("设置字体", self)
        font_action.triggered.connect(self.set_font_with_handler)
        
        color_action = QAction("设置颜色", self)
        color_action.triggered.connect(self.set_color_with_handler)
        
        menu.addActions([font_action, color_action])
        menu.exec_(self.mapToGlobal(pos))

    def set_font_with_handler(self):
        """设置字体 - 使用自定义对话框扩大尺寸"""
        # 创建自定义字体对话框
        font_dialog = QFontDialog(self)
        font_dialog.setCurrentFont(self.label.font())
        
        # 设置对话框尺寸
        font_dialog.resize(800, 600)  # 扩大对话框尺寸
        
        # 设置对话框选项
        font_dialog.setOption(QFontDialog.DontUseNativeDialog, True)  # 禁用原生对话框确保尺寸生效
        font_dialog.setWindowTitle("设置字段字体")
        
        # 添加确定/取消按钮
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
            Qt.Horizontal, font_dialog
        )
        button_box.accepted.connect(font_dialog.accept)
        button_box.rejected.connect(font_dialog.reject)
        
        # 将按钮添加到布局
        layout = font_dialog.layout()
        layout.addWidget(button_box)
        
        # 显示对话框
        if font_dialog.exec_() == QDialog.Accepted:
            font = font_dialog.currentFont()
            # 获取字体系列名称
            font_name = font.family()
            
            # 更新标签字体
            self.label.setFont(font)
            
            # 更新字段数据
            self.field_data.update({
                'font': font_name,
                'size': font.pointSize()
            })
            
    def set_color_with_handler(self):
        color = QColorDialog.getColor(QColor(self.field_data['color']), self)
        if color.isValid():
            self.label.setStyleSheet(f"color: {color.name()};")
            self.field_data['color'] = color.name()
            
    def set_size_with_handler(self):
        size, ok = QInputDialog.getInt(
            self, "字号设置", "字号:", 
            value=self.field_data['size'],
            min=8, max=72
        )
        if ok:
            font = self.label.font()
            font.setPointSize(size)
            self.label.setFont(font)
            self.field_data['size'] = size

    def moveEvent(self, event):
        """限制控件在画布范围内"""
        new_pos = self.pos()
        max_x = self.parent().width() - self.width()
        max_y = self.parent().height() - self.height()
        
        new_pos.setX(max(0, min(new_pos.x(), max_x)))
        new_pos.setY(max(0, min(new_pos.y(), max_y)))
        
        self.move(new_pos)
        self.update_field_data()

    def resizeEvent(self, event):
        """调整大小事件"""
        super().resizeEvent(event)
        self.field_data.update({
            'width': self.width(),
            'height': self.height()
        })

    def toggle_grid(self, visible):
        """切换网格可见性"""
        self.show_grid = visible
        self.update()        


    def create_field(self, field_name, pos):
        """在画布上创建可移动字段控件"""
        # 创建带样式的控件
        field_widget = MovableFieldWidget(self.drop_area, field_name)
        
        # 设置初始位置和样式
        field_widget.move(pos.x()-50, pos.y()-20)  # 微调位置让控件居中显示
        field_widget.setStyleSheet("""
            background: #e3f2fd;
            border: 1px solid #90caf9;
            border-radius: 4px;
            padding: 5px;
        """)
        
        # 添加删除按钮功能
        delete_btn = field_widget.findChild(QPushButton)
        if delete_btn:
            delete_btn.clicked.connect(field_widget.deleteLater)
        
        field_widget.show()
        self.drop_area.update()  # 强制刷新画布


    def set_field_font(self):
        font, ok = QFontDialog.getFont()
        if ok:
            for child in self.findChildren(MovableFieldWidget):
                if child.isSelected():
                    child.label.setFont(font)
                    child.field_data['font'] = font.family()

    def set_field_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            for child in self.findChildren(MovableFieldWidget):
                if child.isSelected():
                    child.label.setStyleSheet(f"color: {color.name()};")
                    child.field_data['color'] = color.name()

    def __del__(self):
        try:
            self.deleteLater()
        except:
            pass

    def deleteLater(self):
        try:
            self.delete_btn.clicked.disconnect()
            self.customContextMenuRequested.disconnect()
            # 清除父对象引用
            self.template_designer = None
            super().deleteLater()
        except:
            pass
# ---------------------------
# 2. 画布容器类
# ---------------------------

class DropArea(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.fields = []
        self.template_designer = parent
        self.background_image = None
        self.background_image_path = None
        self.show_grid = True
        self.grid_size = 20
        
        # 初始化设置
        self.setAcceptDrops(True)
        self.setMinimumSize(800, 600)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)
        
        # 样式设置
        self.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 2px dashed #6c757d;
                border-radius: 5px;
            }
        """)

    def paintEvent(self, event):
        """绘制背景和网格"""
        super().paintEvent(event)
        painter = QPainter(self)
        
        # 绘制背景图片
        if self.background_image:
            painter.drawPixmap(0, 0, self.width(), self.height(), self.background_image)
        
        # 绘制网格
        if self.show_grid:
            painter.setPen(QColor(200, 200, 200, 50))
            for x in range(0, self.width(), self.grid_size):
                painter.drawLine(x, 0, x, self.height())
            for y in range(0, self.height(), self.grid_size):
                painter.drawLine(0, y, self.width(), y)

    def show_context_menu(self, pos):
        """画布右键菜单：处理背景、网格等全局设置"""
        menu = QMenu(self)
        
        # 背景操作
        bg_action = QAction("设置背景图片", self)
        bg_action.triggered.connect(self.set_background_image)
        
        clear_bg_action = QAction("清除背景", self)
        clear_bg_action.triggered.connect(self.clear_background)
        
        # 网格操作
        grid_action = QAction("显示网格", self, checkable=True)
        grid_action.setChecked(self.show_grid)
        grid_action.triggered.connect(lambda: self.toggle_grid(grid_action.isChecked()))
        
        menu.addActions([bg_action, clear_bg_action])
        menu.addSeparator()
        menu.addAction(grid_action)
        menu.exec_(self.mapToGlobal(pos))

    def set_background_image(self):
        """设置背景图片"""
        path, _ = QFileDialog.getOpenFileName(
            self, "选择背景图片", "", 
            "图片文件 (*.png *.jpg *.bmp)"
        )
        if path:
            self.background_image = QPixmap(path)
            self.background_image_path = path
            self.update()

    def clear_background(self):
        """清除背景图片"""
        self.background_image = None
        self.background_image_path = None
        self.update()

    def toggle_grid(self, visible):
        """切换网格显示"""
        self.show_grid = visible
        self.update()

    def dragEnterEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()
            self.setStyleSheet("border-color: #0d6efd;")
            
    def dropEvent(self, event):
        try:
            if event.mimeData().hasText():
                field_name = event.mimeData().text()
                
                # 终极兼容写法（适配所有PyQt5版本和场景）
                if hasattr(event, 'position'):  # 检查是否有position属性
                    pos = event.position().toPoint()  # PyQt5 >= 5.15
                else:
                    pos = event.pos()  # PyQt5 < 5.15 或某些特殊情况
                
                # 调试输出坐标（确认坐标获取正确）
                print(f"Drop position: {pos.x()}, {pos.y()}")  # 调试用
                
                # 创建控件并确保父对象正确
                field_widget = MovableFieldWidget(self, field_name)  # self是DropArea实例
                field_widget.move(pos)
                field_widget.show()
                
                event.acceptProposedAction()
            else:
                event.ignore()
        except Exception as e:
            logger.error(f"拖放失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "错误", f"拖放操作失败: {str(e)}")

    def clear_canvas(self):
        try:
            # 使用列表推导避免遍历时修改集合
            children = [child for child in self.findChildren(MovableFieldWidget)]
            for child in children:
                try:
                    child.setParent(None)
                    child.deleteLater()
                except:
                    continue
            # 强制GC回收
            QApplication.processEvents()  
            self.update()
        except Exception as e:
            logger.error(f"清空画布失败: {str(e)}")

    def dragLeaveEvent(self, event):
        self.setStyleSheet("border: 2px dashed #6c757d;")
        event.accept()       
# ---------------------------
# 3. 主界面类（核心功能）
# ---------------------------
class TemplateDesigner(QDialog):
    def __init__(self, parent=None, db_path=None, selected_ids=None):
        super().__init__(parent)
        
        # ============== 1. 基础属性初始化 ==============
        self.db_path = db_path or get_db_path()
        self.selected_ids = selected_ids or []  # 确保selected_ids是列表
        self.personnel_data = []
        self.current_color = QColor(Qt.black)
        self.current_mode = "standard"
        self.label_mode = False
        self.label_mode_components_initialized = False  # 新增标志位
        if parent:
            parent.init_chinese_font()
        # ============== 2. 字段管理器验证 ==============
        try:
            if not hasattr(field_manager, 'get_field_names'):
                raise RuntimeError("字段管理器未初始化")
            self.fields = field_manager.get_field_names()
        except Exception as e:
            QMessageBox.critical(None, "致命错误", f"字段配置加载失败: {str(e)}")
            sys.exit(1)

    # ============== 3. 核心组件初始化 ==============
        self._init_components()  # 拆分核心初始化逻辑
        
        # ============== 4. 窗口属性设置 ==============
        self.setWindowTitle("高级套打设计器")
        self._setup_window_geometry()
        
        # ============== 5. 调试输出 ==============
        logger.debug(f"设计器初始化完成，字段数：{len(self.fields)}")
        logger.debug(f"选中人员ID：{self.selected_ids}")

    def _init_components(self):
        """初始化核心组件（拆分为独立方法）"""
        # 必须在UI初始化前注册字体
        self.init_font_registry()
        self.init_font_cache()
        
        # 初始化标签模式组件
        self.init_label_mode_components()
        
        # 关键修复：必须在创建子控件前启用拖放
        self.setAcceptDrops(True)  # <--- 添加在这里
        
        # 创建画布区域
        self.drop_area = DropArea(self)
        
        # 初始化UI（确保在最后执行）
        self.init_ui()
        
        # 加载人员数据（必须在UI初始化后）
        self.load_personnel_data()
        
    def _setup_window_geometry(self):
        """设置窗口位置和大小"""
        screen = QApplication.primaryScreen().availableGeometry()
        self.resize(int(screen.width() * 0.7), int(screen.height() * 0.7))
        self.setMinimumSize(800, 500)  # 更合理的默认最小尺寸
        
        # 居中显示
        frame_geo = self.frameGeometry()
        frame_geo.moveCenter(screen.center())
        self.move(frame_geo.topLeft())

    # 在 TemplateDesigner 类的 init_ui 方法中修改
    def init_ui(self):
        """UI初始化"""
        try:
            # 主布局必须最先设置
            self.main_layout = QHBoxLayout(self)  # 注意：直接设置给self
            self.main_layout.setContentsMargins(5, 5, 5, 5)  # 减少边距
            self.main_layout.setSpacing(5)  # 减少组件间距
            
            # 左侧字段面板（宽度调整为更紧凑）
            self.field_panel = self._create_field_panel()
            self.field_panel.setFixedWidth(200)  # 固定宽度，留更多空间给画布
            self.main_layout.addWidget(self.field_panel)
            
            # 右侧设计区域 - 使用垂直布局
            right_container = QWidget()
            self.right_layout = QVBoxLayout(right_container)
            self.right_layout.setContentsMargins(0, 0, 0, 0)
            self.right_layout.setSpacing(5)
            
            # 模式切换工具栏（高度更紧凑）
            mode_toolbar = self.create_mode_toolbar()
            mode_toolbar.setFixedHeight(50)  # 固定工具栏高度
            self.right_layout.addWidget(mode_toolbar)
            
            # 标签模式设置（需要时显示）
            if self.label_mode and hasattr(self, 'label_mode_group'):
                self.label_mode_group.setFixedHeight(120)  # 固定高度
                self.right_layout.addWidget(self.label_mode_group)
            
            # 画布区域 - 使用QFrame提供边框
            canvas_frame = QFrame()
            canvas_frame.setFrameShape(QFrame.StyledPanel)
            canvas_layout = QVBoxLayout(canvas_frame)
            canvas_layout.setContentsMargins(0, 0, 0, 0)
            
            # 画布初始化 - 设置尺寸策略优先扩展
            self.drop_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            canvas_layout.addWidget(self.drop_area)
            
            self.right_layout.addWidget(canvas_frame, 1)  # 拉伸因子设为1，占据剩余空间
            
            # 底部按钮面板（高度更紧凑）
            self._init_bottom_buttons()
            self.btn_panel.setFixedHeight(80)  # 固定按钮面板高度
            self.right_layout.addWidget(self.btn_panel)
            
            self.main_layout.addWidget(right_container, 1)  # 右侧容器占据主要空间
            
            # 设置窗口最小尺寸
            self.setMinimumSize(1000, 700)
            
        except Exception as e:
            logger.error(f"UI初始化失败: {traceback.format_exc()}")
            raise RuntimeError(f"界面创建失败: {str(e)}")
        # ================================================
    def refresh_personnel_data(self):
        """刷新人员数据"""
        self.selected_ids = self.parent().get_selected_personnel_ids()
        self.load_personnel_data()    

    def validate_system_fonts(self):
        """检查系统是否包含宋体字体"""
        available_fonts = QFontDatabase().families()
        if "宋体" not in available_fonts:
            QMessageBox.warning(
                self, "字体缺失",
                "系统缺少宋体字体，界面显示可能异常\n"
                "建议安装中文字体包"
            )
            # 回退到系统默认字体
            self.label.setFont(QFont("Arial", 10))    

    def clear_canvas(self):
        """清空画布上的所有字段控件"""
        if hasattr(self, 'drop_area') and self.drop_area:
            # 安全删除所有字段控件
            for widget in self.drop_area.findChildren(MovableFieldWidget):
                try:
                    widget.setParent(None)
                    widget.deleteLater()
                except:
                    pass
            # 强制刷新界面
            QApplication.processEvents()


    def _create_field_panel(self):
        """创建左侧字段面板"""
        panel = QGroupBox("可用字段")
        layout = QVBoxLayout()
        
        # 确保使用self.fields而不是直接访问field_manager
        if not hasattr(self, 'fields') or not self.fields:
            self.fields = field_manager.get_field_names()  # 重新获取字段列表
        
        scroll = QScrollArea()
        content = QWidget()
        content_layout = QVBoxLayout(content)
        
        # 调试输出字段列表
        print("可用字段:", self.fields)  # 检查字段是否加载
        
        for field in self.fields:
            label = self.DraggableLabel(field)
            content_layout.addWidget(label)
        
        scroll.setWidget(content)
        layout.addWidget(scroll)
        panel.setLayout(layout)
        return panel
    def create_design_panel(self):
        """创建设计面板"""
        panel = QGroupBox("设计画布")
        layout = QVBoxLayout()
        
        # 确保drop_area已初始化
        if not hasattr(self, 'drop_area'):
            self.drop_area = DropArea(self)
            self.drop_area.setMinimumSize(800, 600)
            self.drop_area.setStyleSheet("""
                background-color: white;
                border: 2px dashed #ccc;
            """)
        
        # 模式切换工具栏
        mode_toolbar = self.create_mode_toolbar()
        layout.addWidget(mode_toolbar)
        layout.addWidget(self.drop_area)
        
        panel.setLayout(layout)
        return panel  

    def create_mode_toolbar(self):
        """创建模式切换工具栏"""
        toolbar = QToolBar()
        toolbar.setIconSize(QSize(32, 32))
        
        # 模式切换按钮组
        self.mode_group = QButtonGroup(self)
        
        # 标准模式按钮
        self.btn_standard = QRadioButton("标准模式")
        self.btn_standard.setChecked(True)
        self.btn_standard.toggled.connect(self.handle_mode_change)
        
        # 标签模式按钮
        self.btn_label = QRadioButton("标签模式")
        self.btn_label.toggled.connect(self.handle_mode_change)
        
        # 设置按钮字体大小
        font = QFont()
        font.setPointSize(25)  # +++ 增大工具栏按钮字体 +++
        self.btn_standard.setFont(font)
        self.btn_label.setFont(font)
        
        # 添加到工具栏
        toolbar.addWidget(self.btn_standard)
        toolbar.addWidget(self.btn_label)
        
        # 样式设置
        toolbar.setStyleSheet("""
            QRadioButton {
                padding: 10px 15px; /* +++ 增大内边距 +++ */
                font-size: 14px;    /* +++ 增大字体 +++ */
            }
            QRadioButton::indicator {
                width: 22px;       /* +++ 增大指示器大小 +++ */
                height: 22px;       /* +++ 增大指示器大小 +++ */
            }
        """)
        
        return toolbar
    def init_label_mode_components(self):
        # 添加初始化标志检查
        if hasattr(self, 'label_mode_components_initialized') and self.label_mode_components_initialized:
            return

        self.label_mode_group = QGroupBox("标签模式设置")
        # 移除 WA_DeleteOnClose 属性以避免意外删除
        layout = QFormLayout()
        
        # 标签数量设置
        self.label_count_spin = QSpinBox()
        self.label_count_spin.setRange(1, 20)
        self.label_count_spin.setValue(5)
        layout.addRow(QLabel("每页标签数:"), self.label_count_spin)
        
        # 纸张类型选择
        self.page_size_combo = QComboBox()
        self.page_size_combo.addItems(["A4 (210x297mm)", "Letter (216x279mm)", "自定义"])
        layout.addRow(QLabel("纸张规格:"), self.page_size_combo)
        
        # 边距设置
        self.margin_spin_x = QSpinBox()
        self.margin_spin_x.setRange(0, 50)
        self.margin_spin_x.setValue(10)
        self.margin_spin_y = QSpinBox()
        self.margin_spin_y.setRange(0, 50)
        self.margin_spin_y.setValue(10)
        layout.addRow(QLabel("横向边距(mm):"), self.margin_spin_x)
        layout.addRow(QLabel("纵向边距(mm):"), self.margin_spin_y)
        
        self.label_mode_group.setLayout(layout)
        self.label_mode_group.hide()
        
        # 在类构造函数中初始化这个标志位
        if not hasattr(self, 'label_mode_components_initialized'):
            self.label_mode_components_initialized = False
        self.label_mode_components_initialized = True  # 设置初始化标志

    # ... 其他代码 ...
    def save_template_state(self):
        """保存模板设计器状态"""
        settings = QSettings("MyCompany", "TemplateDesigner")
        # 保存当前模式
        settings.setValue("labelMode", self.label_mode)
        # 保存字段位置和样式
        fields = []
        for widget in self.drop_area.findChildren(MovableFieldWidget):
            fields.append(widget.field_data)
        settings.setValue("fieldData", json.dumps(fields))
        # 保存背景图片路径
        if hasattr(self.drop_area, 'background_image_path'):
            settings.setValue("backgroundImage", self.drop_area.background_image_path)
        # 保存标签模式设置
        if self.label_mode:
            settings.setValue("labelCount", self.label_count_spin.value())
            settings.setValue("marginX", self.margin_spin_x.value())
            settings.setValue("marginY", self.margin_spin_y.value())
            settings.setValue("pageSize", self.page_size_combo.currentIndex())

    def restore_template_state(self):
        """恢复模板设计器状态"""
        settings = QSettings("MyCompany", "TemplateDesigner")
        # 恢复模式
        label_mode = settings.value("labelMode", False, type=bool)
        self.btn_label.setChecked(label_mode)
        self.btn_standard.setChecked(not label_mode)
        self.handle_mode_change()
        # 恢复字段
        try:
            field_data = json.loads(settings.value("fieldData", "[]"))
            for data in field_data:
                self.add_field_to_canvas(
                    data['text'],
                    data['x'],
                    data['y'],
                    data.get('font', 'Arial'),
                    data.get('size', 12),
                    data.get('color', '#000000')
                )
        except Exception as e:
            logger.error(f"恢复字段失败: {str(e)}")
        # 恢复背景
        bg_path = settings.value("backgroundImage", "")
        if bg_path and os.path.exists(bg_path):
            self.drop_area.background_image = QPixmap(bg_path)
            self.drop_area.background_image_path = bg_path
            self.drop_area.update()
        # 恢复标签模式设置
        if label_mode:
            self.label_count_spin.setValue(int(settings.value("labelCount", 5)))
            self.margin_spin_x.setValue(int(settings.value("marginX", 10)))
            self.margin_spin_y.setValue(int(settings.value("marginY", 10)))
            self.page_size_combo.setCurrentIndex(int(settings.value("pageSize", 0)))

    def on_clear_button_clicked(self):  # 假设有个清空按钮
        self.drop_area.clear_canvas() 

    def init_font_registry(self):
        """确保中文字体已注册"""
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # 尝试注册常见中文字体
            font_paths = [
                r"C:\Windows\Fonts\simsun.ttc",  # Windows 宋体
                r"C:\Windows\Fonts\simhei.ttf",   # Windows 黑体
                "/System/Library/Fonts/STHeiti Medium.ttc",  # Mac
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"  # Linux
            ]
            
            # 添加自定义字体路径
            custom_font_path = resource_path("fonts/simsun.ttc")
            if os.path.exists(custom_font_path):
                font_paths.append(custom_font_path)
            
            # 注册找到的第一个可用字体
            for path in font_paths:
                if os.path.exists(path):
                    try:
                        font_name = "chineseFont"
                        pdfmetrics.registerFont(TTFont(font_name, path))
                        logger.info(f"成功注册字体: {path}")
                        return
                    except Exception as e:
                        logger.warning(f"字体注册失败 {path}: {str(e)}")
            
            # 回退到CID字体
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
            logger.warning("使用回退CID字体")
        except Exception as e:
            logger.error(f"字体注册失败: {str(e)}")

    def init_font_cache(self):
        """初始化PDF字体缓存"""
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # Windows系统字体路径
            win_fonts = [
                r"C:\Windows\Fonts\simsun.ttc",  # 宋体
                r"C:\Windows\Fonts\simhei.ttf",   # 黑体
                r"C:\Windows\Fonts\msyh.ttc"      # 微软雅黑
            ]
            
            # Linux系统字体路径
            linux_fonts = [
                "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
            ]
            
            # macOS系统字体路径
            mac_fonts = [
                "/System/Library/Fonts/STHeiti Light.ttc",
                "/System/Library/Fonts/STHeiti Medium.ttc"
            ]
            
            registered = False
            for path in win_fonts + linux_fonts + mac_fonts:
                if os.path.exists(path):
                    try:
                        font_name = os.path.splitext(os.path.basename(path))[0]
                        pdfmetrics.registerFont(TTFont(font_name, path))
                        registered = True
                    except:
                        continue
            
            if not registered:
                QMessageBox.warning(self, "字体警告", 
                    "未找到系统字体，PDF输出可能异常\n"
                    "请安装中文字体文件（如simsun.ttc）到系统字体目录")
        except Exception as e:
            logger.error(f"字体初始化失败: {str(e)}")
            QMessageBox.critical(self, "字体错误", 
                f"字体系统初始化失败: {str(e)}\n"
                "请检查reportlab和Pillow是否安装正确")     

    def _init_bottom_buttons(self):
        """初始化底部按钮面板"""
        self.btn_panel = QWidget()
        self.btn_layout = QVBoxLayout(self.btn_panel)
        
        # 添加一个拉伸因子，将按钮推到下方
        self.btn_layout.addStretch(1)
        
        # 按钮行布局
        button_row = QHBoxLayout()
        
        self.btn_save = QPushButton("保存模板")
        self.btn_save.clicked.connect(self.save_template)
        
        self.btn_load = QPushButton("加载模板")
        self.btn_load.clicked.connect(self.load_template)
        
        self.btn_preview = QPushButton("预览效果")
        self.btn_preview.clicked.connect(self.preview_labels)
        
        self.btn_print = QPushButton("打印")
        self.btn_print.clicked.connect(self.print_labels)
        
        button_row.addWidget(self.btn_save)
        button_row.addWidget(self.btn_load)
        button_row.addWidget(self.btn_preview)
        button_row.addWidget(self.btn_print)
        
        self.btn_layout.addLayout(button_row)
        self.btn_panel.setLayout(self.btn_layout)      



# ================ 核心功能方法 ================
    def generate_filled_pdf(self, output_path, personnel_data):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.colors import HexColor
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics

            # 统一在此处定义页面尺寸
            page_size = A4
            page_width, page_height = page_size

            # ============== 安全获取控件值 ==============
            # 标签模式参数
            items_per_page = 1
            columns = 1
            margin_x = 0
            margin_y = 0

            if self.label_mode:
                # 安全获取每页标签数
                if hasattr(self, 'label_count_spin'):
                    items_per_page = max(1, self.label_count_spin.value())  # 确保最小1个
                else:
                    QMessageBox.warning(self, "配置错误", "标签数量控件未初始化")
                    return False

                # 安全获取边距设置
                if hasattr(self, 'margin_spin_x'):
                    margin_x = self.margin_spin_x.value() * mm
                if hasattr(self, 'margin_spin_y'):
                    margin_y = self.margin_spin_y.value() * mm
                
                # 列数计算（添加容错）
                try:
                    columns = 5 if items_per_page >=5 else max(1, items_per_page)
                    rows_per_page = max(1, math.ceil(items_per_page / columns))
                except ZeroDivisionError:
                    columns = 1
                    rows_per_page = 1

                # 计算单元尺寸（添加最小值保护）
                item_width = max(10, (page_width - 2 * margin_x) / columns)
                item_height = max(10, (page_height - 2 * margin_y) / rows_per_page)
            else:
                # 标准模式参数
                items_per_page = 1
                columns = 1
                rows_per_page = 1
                item_width = page_width 
                item_height = page_height

            # ============== PDF生成核心逻辑 ==============
            c = canvas.Canvas(output_path, pagesize=page_size)

            for idx, person in enumerate(personnel_data or []):  # 添加空值保护
                if idx > 0 and idx % items_per_page == 0:
                    c.showPage()

                # ============== 关键修复：解决字段重叠问题 ==============
                if self.label_mode:
                    # 计算当前标签的位置
                    col = idx % columns
                    row = (idx // columns) % rows_per_page
                    x0 = margin_x + col * item_width
                    y0 = margin_y + row * item_height
                else:
                    # 标准模式位置
                    x0 = 50
                    y0 = page_height - 100 - idx * 200  # 标准模式坐标

                # 绘制背景（添加异常保护）
                if hasattr(self.drop_area, 'background_image_path') and self.drop_area.background_image_path:
                    try:
                        c.drawImage(
                            self.drop_area.background_image_path,
                            x0, y0,
                            width=item_width if self.label_mode else (page_width - 100),
                            height=item_height if self.label_mode else None,
                            preserveAspectRatio=True,
                            mask='auto'
                        )
                    except Exception as e:
                        logger.error(f"背景图片绘制失败: {str(e)}")
                        # 清除无效图片路径防止循环报错
                        self.drop_area.background_image_path = None

                # 绘制字段（添加字段有效性验证）
                for widget in self.drop_area.findChildren(MovableFieldWidget):
                    field = widget.field_data
                    try:
                        value = str(person.get(field.get('text', ''), ''))
                        
                        # 标签模式下计算字段在标签内的相对位置
                        if self.label_mode:
                            field_x = x0 + field.get('x', 0)
                            field_y = y0 + field.get('y', 0)
                        else:
                            # 标准模式下直接使用设计时的位置
                            field_x = field.get('x', 50)
                            field_y = page_height - field.get('y', 100)  # 翻转Y轴
                        
                        # 样式设置
                        font_name = field.get('font', 'chineseFont')
                        if font_name not in pdfmetrics.getRegisteredFontNames():
                            font_name = 'chineseFont'  # 强制使用注册字体
                        font_size = field.get('size', 12)
                        font_color = field.get('color', '#000000')
                        
                        # 创建文本对象
                        text = c.beginText(field_x, field_y)
                        text.setFont(font_name, font_size)
                        text.setFillColor(HexColor(font_color))
                        
                        # 处理多行文本（兼容Windows/Linux/macOS换行符）
                        lines = value.replace('\r\n', '\n').split('\n')
                        
                        for i, line in enumerate(lines):
                            # 非首行需要换行（行间距为字号的1.2倍）
                            if i > 0:
                                text.moveCursor(0, -font_size * 1.2)
                            
                            # 直接使用Unicode文本
                            text.textLine(line.strip())
                        
                        # 绘制文本到Canvas
                        c.drawText(text)
                        
                    except Exception as e:
                        logger.error(f"文本渲染失败 | 字段：{field.get('text','')} | 错误：{str(e)}")
                        # 在调试模式下显示错误信息
                        c.setFont("Helvetica", 8)
                        c.setFillColor(HexColor("#FF0000"))
                        c.drawString(50, 50, f"Render Error: {field.get('text','')}")

            c.save()
            return True

        except Exception as e:
            logger.error(f"PDF生成失败: {traceback.format_exc()}")
            QMessageBox.critical(
                self, "生成错误",
                f"PDF生成失败: {str(e)}\n"
                "常见原因：\n"
                "1. 控件参数异常\n"
                "2. 字体配置错误\n"
                "3. 坐标计算越界"
            )
            return False
    # 在预览方法中添加异常捕获
    def preview_labels(self):
        try:
            self.refresh_personnel_data()
            # 添加前置检查
            if not hasattr(self, 'personnel_data') or not self.personnel_data:
                self.load_personnel_data()
                if not self.personnel_data:
                    QMessageBox.warning(self, "数据错误", "没有可用的打印数据")
                    return

            # 生成临时文件路径
            temp_path = os.path.join(
                tempfile.gettempdir(), 
                f"preview_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
            
            # 强制刷新界面状态
            QApplication.processEvents()
            
            # 生成PDF
            if self.generate_filled_pdf(temp_path, self.personnel_data):
                QDesktopServices.openUrl(QUrl.fromLocalFile(temp_path))
        except Exception as e:
            logger.error(f"预览失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "生成错误", 
                f"PDF生成失败: {str(e)}\n"
                "请检查：\n"
                "1. 所有字段位置是否有效\n"
                "2. 图片路径是否正确\n"
                "3. 是否选择了打印人员")

    def print_labels(self):
        try:
            self.refresh_personnel_data()
            # 1. 前置检查
            if not hasattr(self, 'personnel_data') or not self.personnel_data:
                self.load_personnel_data()
                
            if not self.personnel_data:
                QMessageBox.warning(self, "数据错误", "没有可打印的人员数据")
                return

            # 2. 创建带时间戳的临时文件
            temp_filename = f"print_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            pdf_path = os.path.join(tempfile.gettempdir(), temp_filename)

            # 3. 生成PDF（包含字体注册和背景处理）
            try:
                # 增强字体注册
                self.init_font_registry()
                
                # 强制刷新画布状态
                self.drop_area.update()
                
                if not self.generate_filled_pdf(pdf_path, self.personnel_data):
                    raise RuntimeError("PDF生成失败")
                    
                if not os.path.exists(pdf_path):
                    raise FileNotFoundError("PDF文件生成失败")

            except Exception as e:
                logger.error(f"PDF生成失败: {traceback.format_exc()}")
                QMessageBox.critical(
                    self, "生成错误",
                    f"无法生成打印文件：{str(e)}\n"
                    "常见原因：\n"
                    "1. 字体配置错误\n"
                    "2. 图片路径无效\n"
                    "3. 模板字段异常"
                )
                return

            # 4. 打印流程
            try:
                printer = QPrinter(QPrinter.HighResolution)
                print_dialog = QPrintDialog(printer, self)
                
                if print_dialog.exec_() == QDialog.Accepted:
                    # 跨平台打印处理
                    if sys.platform == "win32":
                        os.startfile(pdf_path, "print")
                    elif sys.platform == "darwin":
                        subprocess.run(["lp", "-o", "fit-to-page", pdf_path])
                    else:  # Linux
                        subprocess.run(["lp", "-o", "fit-to-page", pdf_path])
                    
                    # 延迟清理临时文件（30秒后）
                    QTimer.singleShot(30000, lambda: self.clean_temp_file(pdf_path))
                    
            except Exception as e:
                logger.error(f"打印操作失败: {traceback.format_exc()}")
                QMessageBox.critical(
                    self, "打印错误",
                    f"打印指令发送失败：{str(e)}\n"
                    "建议操作：\n"
                    "1. 检查打印机是否在线\n"
                    "2. 尝试手动打印生成的文件：\n"
                    f"{pdf_path}"
                )

        except Exception as e:
            logger.error(f"打印流程异常: {traceback.format_exc()}")
            QMessageBox.critical(
                self, "系统错误",
                f"打印流程发生意外错误：{str(e)}\n"
                "请联系技术支持"
            )

    def clean_temp_file(self, path):
        """安全清理临时文件"""
        try:
            if os.path.exists(path):
                os.remove(path)
                logger.info(f"已清理临时文件: {path}")
        except Exception as e:
            logger.warning(f"临时文件清理失败: {str(e)}")

    def export_pdf(self):
        try:
            selected_ids = self.parent().get_selected_personnel_ids()
            if not selected_ids:
                QMessageBox.warning(self, "警告", "请先在主界面选择人员")
                return

            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            pdf_path, _ = QFileDialog.getSaveFileName(
                self, "导出PDF文件", desktop, "PDF文件 (*.pdf)"
            )
            if not pdf_path:
                return

            self.print_document(pdf_path, selected_ids)

            if os.path.exists(pdf_path):
                QMessageBox.information(self, "成功", f"PDF已导出到：\n{pdf_path}")
                QDesktopServices.openUrl(QUrl.fromLocalFile(pdf_path))
            else:
                QMessageBox.critical(self, "错误", "PDF生成失败")

        except Exception as e:
            logger.error(f"PDF导出失败: {traceback.format_exc()}")
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")        

    def save_template(self):
        template = {'fields': []}
        # 修正子控件查找方式
        for child in self.drop_area.findChildren(MovableFieldWidget):  # 添加self限定类
            if hasattr(child, 'field_data'):
                template['fields'].append(child.field_data)
        
        path, _ = QFileDialog.getSaveFileName(self, "保存模板", "", "JSON文件 (*.json)")
        if path:
            try:
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(template, f, indent=2, ensure_ascii=False)
                QMessageBox.information(self, "成功", "模板保存成功！")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"保存失败：{str(e)}")

    def load_template(self):
        path, _ = QFileDialog.getOpenFileName(self, "加载模板", "", "模板文件 (*.json)")
        if path:
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    template = json.load(f)
                    
                # 使用新的 clear_canvas 方法
                self.clear_canvas()
                
                for field in template.get('fields', []):
                    self.add_field_to_canvas(
                        field.get('text', ''),
                        field.get('x', 0),
                        field.get('y', 0),
                        field.get('font', 'Arial'),
                        field.get('size', 12),
                        field.get('color', '#000000')
                    )
                QMessageBox.information(self, "成功", "模板加载成功！")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"加载失败: {str(e)}")
# ================ 事件处理 ================
# 在 handle_mode_change 方法中添加保护机制
    def handle_mode_change(self):
        try:
            self.label_mode = self.btn_label.isChecked()
            
            # 保证只初始化一次
            if not hasattr(self, 'label_mode_group'):
                self.init_label_mode_components()
                self.right_layout.insertWidget(1, self.label_mode_group)
                
            # 使用show/hide代替setVisible保持对象存活
            self.label_mode_group.setVisible(self.label_mode)
            
            # 清除画布时安全删除子对象
            self.drop_area.clear_canvas()
                
        except Exception as e:
            logger.error(f"模式切换失败: {traceback.format_exc()}")
    def update_ui_for_mode(self):
        """根据当前模式更新UI"""
        if hasattr(self, 'label_mode_group'):
            self.label_mode_group.setVisible(self.label_mode)
        
        if hasattr(self, 'btn_panel'):
            self.btn_panel.setVisible(True)
        
        if hasattr(self, 'drop_area'):
            self.drop_area.setStyleSheet(
                "background-color: #F0FFF0;" if self.label_mode else "background-color: white;"
            )
# ================ 数据管理 ================
    # 修改 TemplateDesigner 类的 load_personnel_data 方法
    def load_personnel_data(self):
        """增强版数据加载方法"""
        if not self.selected_ids:
            QMessageBox.warning(self, "警告", "未选择任何人员！")
            return
        try:
            if not self.selected_ids:
                logger.warning("未接收到selected_ids参数")
                return
                
            logger.debug(f"正在加载人员数据，IDs: {self.selected_ids}")
            
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                placeholders = ','.join(['?'] * len(self.selected_ids))
                query = f"""
                    SELECT * FROM personnel 
                    WHERE 身份证号 IN ({placeholders})
                    ORDER BY 姓名
                """
                cursor.execute(query, self.selected_ids)
                self.personnel_data = [dict(row) for row in cursor.fetchall()]
                
            logger.info(f"成功加载 {len(self.personnel_data)} 条人员数据")
            
        except sqlite3.Error as e:
            logger.error(f"数据库查询失败: {str(e)}")
            QMessageBox.critical(
                self, "数据库错误",
                f"无法加载人员数据:\n{str(e)}\n"
                f"请检查数据库文件: {self.db_path}"
            )
        except Exception as e:
            logger.error(f"数据加载异常: {traceback.format_exc()}")
            raise

    def add_field_to_canvas(self, text, x=0, y=0, font='Arial', size=12, color='#000000'):
        field_widget = MovableFieldWidget(self.drop_area, text)
        field_widget.move(x, y)
        # 设置字体和颜色
        font_obj = QFont(font, size)
        field_widget.label.setFont(font_obj)
        field_widget.label.setStyleSheet(f"color: {color};")
        # 更新字段数据
        field_widget.field_data.update({
            'font': font,
            'size': size,
            'color': color,
            'x': x,
            'y': y,
            'width': field_widget.width(),
            'height': field_widget.height()
        })
        field_widget.show()
    # ================ 嵌套类 ================
    class DraggableLabel(QLabel):
        def __init__(self, text, parent=None):
            super().__init__(text, parent)
            self.setStyleSheet("""
                background-color: #e0e0e0;
                border: 1px solid #a0a0a0;
                padding: 5px;
                font-size: 14px;
                margin: 2px;
            """)
            self.setFixedSize(120, 40)
            self.setCursor(Qt.OpenHandCursor)
            
        def mousePressEvent(self, event):
            if event.button() == Qt.LeftButton:
                # 创建拖拽对象
                drag = QDrag(self)
                mime = QMimeData()
                mime.setText(self.text())
                drag.setMimeData(mime)
             
                
                # 设置拖拽时的缩略图
                pixmap = QPixmap(self.size())
                self.render(pixmap)
                drag.setPixmap(pixmap)
                drag.setHotSpot(event.pos())
                
                # 执行拖拽
                drag.exec_(Qt.CopyAction)

        def mouseMoveEvent(self, event):
            if self.dragging and event.buttons() & Qt.LeftButton:
                new_pos = self.mapToParent(event.pos() - self.offset)
                
                # 限制移动范围在画布内
                new_pos.setX(max(0, min(new_pos.x(), self.parent().width() - self.width())))
                new_pos.setY(max(0, min(new_pos.y(), self.parent().height() - self.height())))
                
                self.move(new_pos)
                self.update_field_data()
            super().mouseMoveEvent(event)

        def mouseReleaseEvent(self, event):
            self.dragging = False
            self.setCursor(Qt.OpenHandCursor)
            super().mouseReleaseEvent(event)

        def deleteLater(self):
            try:
                # 安全断开所有信号
                self.delete_btn.clicked.disconnect()
                self.customContextMenuRequested.disconnect()
                # 显式移除控件
                if self.parent():
                    self.parent().layout().removeWidget(self)
                super().deleteLater()
            except Exception as e:
                logger.error(f"删除控件失败: {str(e)}")
        
        


def main():
    # 1. 检查依赖
    if not check_dependencies():
        QMessageBox.critical(
            None, 
            "环境错误", 
            "缺少打印组件依赖，请安装：\n\npip install reportlab Pillow"
        )
        sys.exit(1)
    
    # 2. 创建QApplication实例
    
    # 3. 初始化主窗口
    window = ArchiveManager()  # 或你的主窗口类名
    window.show()
    
    # 4. 运行主循环
    sys.exit(app.exec_())

        # --------------------------- 运行入口 ---------------------------
# --------------------------- 运行入口 ---------------------------
# ======================= 主程序入口 =======================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")  # 使用Fusion风格确保跨平台一致性
    
    # 初始化数据库
    try:
        init_database()
    except Exception as e:
        QMessageBox.critical(None, "数据库错误", f"数据库初始化失败: {str(e)}")
        sys.exit(1)
    
    # 创建并显示主窗口
    main_window = ArchiveManager()
    main_window.show()
    
    # 确保在显示后恢复状态
    QTimer.singleShot(100, main_window.restore_state)
    
    sys.exit(app.exec_())